from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
import re
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import xlsxwriter
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook

insurance = Blueprint('insurance', __name__)

@insurance.route('/')
def index():
    return render_template('test.html')

@insurance.route('/insurance-upload', methods=['POST'])
def upload_files():

    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\ITEM DESCRIPTION WITH CATEGORY (map).xlsx'
    # Check if an allocation file was provided by the user
    allocation_file = request.files.get('allocation')
    if allocation_file:
        # If the user provided an allocation file, delete the existing default file
        if os.path.exists(default_allocation_path):
            os.remove(default_allocation_path)  # Delete the existing default allocation file
        allocation_file.save(default_allocation_path)  # Save the new allocation file
        category_allocation = pd.read_excel(allocation_file)
    else:
        # If no allocation file is provided, use the default allocation file
        category_allocation = pd.read_excel(default_allocation_path)



# Read the uploaded Excel files
    sales = request.files['sales']
    sales_report_file24_25 = pd.read_excel(sales)


    category_allocation['Allocation'] = category_allocation['Allocation'].str.upper()
    category_allocation.loc[category_allocation['Allocation'].str.contains(r'BULK', case=False),'Allocation']= 'PKT'
    category_allocation.loc[category_allocation['Allocation'].str.contains(r'TEA INSTANT', case=False),'Allocation']= 'TEABAG'
    category_allocation['Allocation'].unique()

    ##################################################################################################################################################3

    sales_report_file24_25['Invoice Date'] = pd.to_datetime(sales_report_file24_25['Phyto Date'])
    sales_report_file24_25['Month'] = sales_report_file24_25['Phyto Date'].dt.strftime('%B')
    sales_report_file24_25 = sales_report_file24_25[~sales_report_file24_25['Item Description'].str.contains('rawtea', case=False, na=False)]

    #FPR  DOMESTIC  and  EXPORT...................................................

    Branch = ['KOL DOM', 'CBE EXP', 'CBEUS', 'KOL EOU','CBE DOM']

    filtered_Branch =sales_report_file24_25[sales_report_file24_25['Branch'].isin(Branch)]

    GL_Description = [
        '3000040000 - Sales Tea Export',
        '3000100000 - Sales Tea Domestic',
        '3000050000 - Sales Packing Material Export']

    filter_GL =filtered_Branch[filtered_Branch['G/L Ac-Description'].isin(GL_Description)]

    filter_GL.loc[filter_GL['Department'].str.contains(r'RED', case=False), 'Department'] = 'Total Red Sales'
    filter_GL.loc[filter_GL['Department'].str.contains(r'YELLOW', case=False), 'Department'] = 'Total Yellow Sales'

    Customer_No_russia= ['CUST00482']
    exclude_russia = filter_GL[~filter_GL['Customer No'].isin(Customer_No_russia)]

    include_dpt = ['Horeca Sales','Online Sales','Private Lable Sales','TEA ME Sales','Total Red Sales','Total Yellow Sales',
                'Export Africa Sales','Export AUS/EUR Sales','Other Export Sales', 'Tea Instant']

    filtered_dept=exclude_russia[exclude_russia['Department'].isin(include_dpt)]

    filtered_dept['Segment'] = filtered_dept['Department'].apply(lambda x: 'Domestic Sales Figure' if x in ['Online Sales', 'Horeca Sales', 'Private Lable Sales', 'TEA ME Sales', 'Total Red Sales', 'Total Yellow Sales'] else 'Export Sales Figure')

    # filtered_dept.to_excel("D:\\1MJPIL_COIMBATORE\\Insurance\\testing_double_data.xlsx")

    sales_table = pd.pivot_table(filtered_dept, index=['Month'],columns='Segment', 
                                values='Sale Amount Actual',aggfunc='sum', fill_value=0)
    month_order = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March']
    sales_table = sales_table.reindex(month_order, axis=0)


    #FOR  RED ...............................................................

    Branch = ['RED BIH', 'RED JKD','RED MP', 'PKT MAH','RED UP', 'RED HP', 'RED PUN','RED UK', 'RED RAJ']

    red_Branch =sales_report_file24_25[sales_report_file24_25['Branch'].isin(Branch)]

    red_GL_Description = ['3000100000 - Sales Tea Domestic']

    red_GL =red_Branch[red_Branch['G/L Ac-Description'].isin(red_GL_Description)]

    exclude_dpt = ['Depot Transfer']

    filtered_RED_df=red_GL[~(red_GL['Department'].isin(exclude_dpt))]

    filtered_RED_df['Branch'].unique()
    filtered_RED_df['Segment'] = 'RED Sales Figure'

    RED_table = pd.pivot_table(filtered_RED_df, index=['Month'],columns='Segment', 
                                values='Sale Amount Actual',aggfunc='sum', fill_value=0)
    RED_table = RED_table.reindex(month_order, axis=0)

    sale_red_table = pd.concat([sales_table,RED_table], axis = 1)

    sale_red_table['Insured Value - (Sales Figure + 10%)'] = (sale_red_table['Domestic Sales Figure'] + sale_red_table['Export Sales Figure'] + sale_red_table['RED Sales Figure'])*1.10
    sale_red_table['Domestic Premium'] = ((sale_red_table['Domestic Sales Figure']* 1.10) * 0.02) / 100
    sale_red_table['Export Premium'] = ((sale_red_table['Export Sales Figure']* 1.10) * 0.02) / 100
    sale_red_table['RED Premium'] = ((sale_red_table['RED Sales Figure']* 1.10) * 0.02) / 100
    sale_red_table['Total Premium'] = sale_red_table['Domestic Premium'] + sale_red_table['Export Premium'] + sale_red_table['RED Premium']


    ####################################################################################################################################################

    category_filter = category_allocation.drop_duplicates(subset=['Item Description','Allocation'])
    merged_df= pd.merge(filtered_dept, category_filter, on='Item Description', how='left')

    merged_df.loc[merged_df['Branch'] == 'CBE DOM', 'Allocation'] = ''


    #######################################################################################################################################################

    data_domestic = merged_df[(merged_df['Segment'] == 'Domestic Sales Figure')] 

    data_domestic_group = data_domestic.groupby(['Month','Allocation','Branch']).agg({'Sale Amount Actual': 'sum'}).reset_index()

    data_domestic_group['Month'] = pd.Categorical(data_domestic_group['Month'], categories=month_order, ordered=True)

    branch_order = ['KOL DOM', 'CBE DOM']

    data_domestic_group['Branch'] = pd.Categorical(data_domestic_group['Branch'], categories=branch_order, ordered=True)

    data_domestic_group = data_domestic_group.sort_values(by=['Month', 'Branch'])

    data_domestic_group = data_domestic_group.rename(columns={'Sale Amount Actual': '100% INR VALUE'})
    data_domestic_group['110% INR VALUE'] = data_domestic_group['100% INR VALUE']*1.10
    data_domestic_group['Premium'] = (( data_domestic_group['100% INR VALUE']* 1.10) * 0.02) / 100



    data_export = merged_df[(merged_df['Segment'] == 'Export Sales Figure')] 

    data_export['Invoice No'] = data_export['Invoice No'].replace({np.nan: '-', ' ': '-'})

    Customer_No= ['CUST00482']
    data_export = data_export[~data_export['Customer No'].isin(Customer_No)]

    data_export_group = data_export.groupby(['Month','Branch','Invoice No']).agg({'Sale Amount Actual': 'sum'}).reset_index()

    data_export_group['Month'] = pd.Categorical(data_export_group['Month'], categories=month_order, ordered=True)

    branch_order = ['KOL EOU','CBE EXP', 'CBEUS']

    data_export_group['Branch'] = pd.Categorical(data_export_group['Branch'], categories=branch_order, ordered=True)
    data_export_group = data_export_group.sort_values(by=['Month', 'Branch'])
    data_export_group = data_export_group.rename(columns={'Sale Amount Actual': '100% INR VALUE'})
    data_export_group['110% INR VALUE'] = data_export_group['100% INR VALUE']*1.10
    data_export_group['Premium'] = (( data_export_group['100% INR VALUE']* 1.10) * 0.02) / 100


    ##################################################################################################################################################

    filtered_dept_russia=filter_GL[filter_GL['Department'].isin(include_dpt)]

    filtered_dept_russia['Segment'] =filtered_dept_russia['Department'].apply(lambda x: 'Domestic Sales Figure' if x in ['Online Sales', 'Horeca Sales', 'Private Lable Sales', 'TEA ME Sales', 'Total Red Sales', 'Total Yellow Sales'] else 'Export Sales Figure')

    category_filter = category_allocation.drop_duplicates(subset=['Item Description','Allocation'])
    merged_df= pd.merge(filtered_dept_russia, category_filter, on='Item Description', how='left')

    merged_df.loc[merged_df['Branch'] == 'CBE DOM', 'Allocation'] = ''

    Customer_No= ['CUST00482']
    russia_export = merged_df[merged_df['Customer No']== 'CUST00482']

    russia_export_group = russia_export.groupby(['Month','Branch','Invoice No']).agg({'Sale Amount Actual': 'sum'}).reset_index()

    russia_export_group['Month'] = pd.Categorical(russia_export_group['Month'], categories=month_order, ordered=True)

    branch_order = ['KOL EOU','CBE EXP', 'CBEUS']

    russia_export_group['Branch'] = pd.Categorical(russia_export_group['Branch'], categories=branch_order, ordered=True)
    russia_export_group = russia_export_group.sort_values(by=['Month', 'Branch'])
    russia_export_group = russia_export_group.rename(columns={'Sale Amount Actual': '100% INR VALUE'})
    russia_export_group['110% INR VALUE'] = russia_export_group['100% INR VALUE']*1.10
    russia_export_group['Premium'] = ((russia_export_group['100% INR VALUE']* 1.10) * 0.1) / 100


    ######################################################################################################################################################


    russia_table = pd.pivot_table(russia_export, index=['Month'],columns='Segment', 
                                values='Sale Amount Actual',aggfunc='sum', fill_value=0)

    russia_table = russia_table.reindex(month_order, axis=0)
    russia_table = russia_table.rename(columns={'Export Sales Figure': 'Russia Sales Figure'})
    russia_table ['Insured Value - (Sales Figure + 10%)'] = (russia_table['Russia Sales Figure'])*1.10
    russia_table ['Russia Premium'] = ((russia_table ['Russia Sales Figure']* 1.10) * 0.1) / 100


    ####################################################################################################################################################

    Red_group = filtered_RED_df.groupby(['Month', 'Branch', 'Department']).agg({'Sale Amount Actual': 'sum'}).reset_index()

    Red_group['Month'] = pd.Categorical(Red_group['Month'], categories=month_order, ordered=True)

    Red_group = Red_group.sort_values(by=['Month'])

    Red_group = Red_group.rename(columns={'Sale Amount Actual': '100% INR VALUE'})
    Red_group['110% INR VALUE'] = Red_group['100% INR VALUE']*1.10
    Red_group['Premium'] = ((Red_group['100% INR VALUE']* 1.10) * 0.02) / 100

    #####################################################################################################################################################

    error_df = merged_df[(merged_df['Branch'] == 'KOL DOM') & (merged_df['Allocation'].isna())]
    error_file = error_df[['Item Description','Allocation']]
    error_file = error_file.drop_duplicates()


    # Create an in-memory output file
    output_file = io.BytesIO()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        sale_red_table.to_excel(writer, sheet_name='exclduRuss_Summary', index=True)
        russia_table.to_excel(writer, sheet_name='Russia_Summary', index=True)
        data_domestic_group.to_excel(writer, sheet_name='Domestic', index=False)
        data_export_group.to_excel(writer, sheet_name='export invoice_excluding Russia', index=False)
        russia_export_group.to_excel(writer, sheet_name='Invoice Russia', index=False)
        merged_df.to_excel(writer, sheet_name='Raw_File', index=False)
        Red_group.to_excel(writer, sheet_name='RED_Branch', index=False)
        filtered_RED_df.to_excel(writer, sheet_name='RED_RawFile', index=False)
        error_file.to_excel(writer, sheet_name='ERROR_Rows', index=False)

    # Save the Excel file to the in-memory object
    output_file.seek(0)  # Reset the buffer position to the beginning of the file

    # Save to a temporary file to apply formatting
    with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp.write(output_file.getvalue())
        tmp.seek(0)

        # Load the temporary file with openpyxl
        wb = load_workbook(tmp.name)

        # Apply formatting
        sheets_to_format = [
            'exclduRuss_Summary', 'Russia_Summary', 'Domestic', 
            'export invoice_excluding Russia', 'Invoice Russia', 'RED_Branch']
        bold_font = Font(bold=True)
        border_style = Border(left=Side(border_style='thin'), 
                            right=Side(border_style='thin'),
                            top=Side(border_style='thin'), 
                            bottom=Side(border_style='thin'))

        def apply_comma_format(sheet):
            for row in sheet.iter_rows(min_row=2, min_col=1):
                for cell in row:
                    cell.number_format = '#,##0'
                    cell.border = border_style

        def autofit_columns(sheet):
            for column in sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2)
                sheet.column_dimensions[column_letter].width = adjusted_width

        for sheet_name in sheets_to_format:
            sheet = wb[sheet_name]
            for cell in sheet[1]:
                cell.font = bold_font
                cell.border = border_style
            apply_comma_format(sheet)
            autofit_columns(sheet)

        # Save the modified Excel file back to the temporary file
        wb.save(tmp.name)

        # Read the temporary file back into the in-memory BytesIO object
        tmp.seek(0)
        output_file = io.BytesIO(tmp.read())

    # Prepare the in-memory file for download
    output_file.seek(0)  # Reset the buffer before sending for download
    current_app.config['EXCEL_FILE'] = {
        'name': 'Insurance_MIS_Report.xlsx',
        'data': output_file.getvalue()}


    # Redirect to the processing page
    return redirect(url_for('processing'))

@insurance.route('/processing')
def processing():
    return render_template('processing.html')

@insurance.route('/download')
def download_report():
# Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])

@insurance.route('/insurance-master', methods=['GET'])
def insurance_master():
    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\ITEM DESCRIPTION WITH CATEGORY (map).xlsx'
    return send_file(default_allocation_path, as_attachment=True)

if __name__ == '__main__':
    insurance.run(debug=True)