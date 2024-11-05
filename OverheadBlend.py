from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
import re
import numpy as np
import openpyxl as px
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import xlsxwriter
from tempfile import NamedTemporaryFile
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

blendovr = Blueprint('blendovr', __name__)

# Define upload folder path
UPLOAD_FOLDER = r'D:\2MJIPL_COIMBATORE\BLEND DATA FOLDER'
if not os.path.exists(UPLOAD_FOLDER):
    os.makedirs(UPLOAD_FOLDER)

# Function to clean file paths
def clean_path(path):
    return path.strip().replace('\u202a', '')


@blendovr.route("/", methods=["GET"])
def index():
    return render_template("text.html")

@blendovr.route('/blend-upload', methods=['POST'])
def upload_files():

    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\Blend Mapping Master.xlsx'
    # Check if an allocation file was provided by the user
    allocation_file = request.files.get('allocation')
    if allocation_file:
        # If the user provided an allocation file, delete the existing default file
        if os.path.exists(default_allocation_path):
            os.remove(default_allocation_path)  # Delete the existing default allocation file
        allocation_file.save(default_allocation_path)  # Save the new allocation file
        allocation_raw = pd.read_excel(allocation_file, sheet_name ='Expenses_Master' )
        CBE_allocation = pd.read_excel(allocation_file, sheet_name = 'CBE_Blend')
        blend_map = pd.read_excel(allocation_file, sheet_name ='DOM_EOU_Blend' )
    else:
        # If no allocation file is provided, use the default allocation file
        allocation_raw = pd.read_excel(default_allocation_path, sheet_name ='Expenses_Master' )
        CBE_allocation = pd.read_excel(default_allocation_path, sheet_name = 'CBE_Blend')
        blend_map = pd.read_excel(default_allocation_path, sheet_name ='DOM_EOU_Blend' )


    # Path to the default last year sales file
    default_lastyr_path = r'D:\2MJIPL_COIMBATORE\Last Year Blend Overhead.xlsx'
    # Check if a last year sales file was provided by the user
    lastyr_file = request.files.get('rately')
    if lastyr_file:
        # If the user provided a last year sales, delete the existing default file
        if os.path.exists(default_lastyr_path):
            os.remove(default_lastyr_path)  # Delete the existing default_lastyr_path
        lastyr_file.save(default_lastyr_path)  # Save the new last year sales file
        rate_last_yr = pd.read_excel(lastyr_file,sheet_name = 'Rate_per_Kg')
    else:
        # If no allocation file is provided, use the default allocation file
        rate_last_yr = pd.read_excel(default_lastyr_path,sheet_name = 'Rate_per_Kg')          


    # chaimaster = request.files['chaimaster']
    # domeoumaster = request.files['domeoumaster']
    # cbemaster =request.files['cbemaster']
    dptfile = request.files['dptfile']
    dpt_raw = pd.read_excel(dptfile)
    # rately = request.files['rately']
 # Retrieve uploaded files from form
    kol_dom_file = request.files['kol_dom']
    kol_eou_file = request.files['kol_eou']
    cbe_file = request.files['cbe']

    # Save uploaded files to the server
    if kol_dom_file:
        kol_dom_filename = secure_filename(kol_dom_file.filename)
        kol_dom_path = os.path.join(UPLOAD_FOLDER, "KOL DOM", kol_dom_filename)
        os.makedirs(os.path.dirname(kol_dom_path), exist_ok=True)
        kol_dom_file.save(kol_dom_path)

    if kol_eou_file:
        kol_eou_filename = secure_filename(kol_eou_file.filename)
        kol_eou_path = os.path.join(UPLOAD_FOLDER, "KOL EOU", kol_eou_filename)
        os.makedirs(os.path.dirname(kol_eou_path), exist_ok=True)
        kol_eou_file.save(kol_eou_path)

    if cbe_file:
        cbe_filename = secure_filename(cbe_file.filename)
        cbe_path = os.path.join(UPLOAD_FOLDER, "CBE", cbe_filename)
        os.makedirs(os.path.dirname(cbe_path), exist_ok=True)
        cbe_file.save(cbe_path)

    # Process the saved files and combine into a DataFrame
    file_paths = {
        "KOL DOM": os.path.join(UPLOAD_FOLDER, "KOL DOM"),
        "KOL EOU": os.path.join(UPLOAD_FOLDER, "KOL EOU"),
        "CBE": os.path.join(UPLOAD_FOLDER, "CBE")
    }

    month_array =['April','May','June', 'July','August','September','October','November',
              'December','January','February', 'March']

    CBE_allocation = CBE_allocation[['Branch', 'Departments','Account No', 'Allocation']]

    dpt_raw['Voucher Date'] = pd.to_datetime(dpt_raw['Voucher Date'])
    dpt_raw['Month_Name'] = dpt_raw['Voucher Date'].dt.strftime('%B')

    ######################################################################################################################

    filter_for_DOM_EOU =dpt_raw[
        ((dpt_raw['Branch'] == 'Kolkata DOM Branch') | (dpt_raw['Branch'] == 'Kolkata EOU Branch')|(dpt_raw['Branch'] == 'Guwahati Branch')|(dpt_raw['Branch'] == 'Kolkata Common Branch')) &
        ((dpt_raw['Departments'] == 'Tea Procurement') | (dpt_raw['Departments'] == 'Specialty'))]


    filter_for_CBE =dpt_raw[
        ((dpt_raw['Departments'] == 'Tea Procurement') | (dpt_raw['Departments'] == 'Specialty')|
        (dpt_raw['Departments'] == 'Tea Instant')) &
        
        ((dpt_raw['Branch'] == 'CBE Print') |
        (dpt_raw['Branch'] == 'CBEUS') |
        (dpt_raw['Branch'] == 'Coimbatore Common Branch') |
        (dpt_raw['Branch'] == 'Coimbatore Domestic Branch') |
        (dpt_raw['Branch'] == 'Coimbatore Export Branch'))]


    ####################################################################################################################

    allocation_df = allocation_raw.drop_duplicates(subset=['Branch', 'Departments','Account No','Allocation'])

    CBE_allocation = CBE_allocation.drop_duplicates(subset=['Branch', 'Departments', 'Account No','Allocation'])

    allocation_df['Account No'] = allocation_df['Account No'].fillna(0)
    allocation_df['Account No'] = allocation_df['Account No'].astype('int64')
    filter_for_DOM_EOU['Account No'] = filter_for_DOM_EOU['Account No'].astype('int64')

    merged_DOM_EOU = pd.merge(filter_for_DOM_EOU,allocation_df, on=['Branch', 'Departments', 'Account No'], how='left')
    blank_check =merged_DOM_EOU.copy()

    CBE_allocation['Account No'] = CBE_allocation['Account No'].fillna(0)
    CBE_allocation['Account No'] = CBE_allocation['Account No'].astype('int64')
    filter_for_CBE['Account No'] = filter_for_CBE['Account No'].astype('int64')

    merged_CBE = pd.merge(filter_for_CBE,CBE_allocation, on=['Branch', 'Departments','Account No'], how='left')
    blank_cbe = merged_CBE.copy()
    merged_CBE['Branch'] = 'CBE'

    merged_DOM_EOU['Branch'] = merged_DOM_EOU['Branch'].replace('Kolkata Common Branch', 'Kolkata EOU Branch')
    merged_DOM_EOU['Branch'] = merged_DOM_EOU['Branch'].replace('Guwahati Branch', 'Kolkata DOM Branch')

    required_allocation_DOM_EOU = ['DOM- Tea Procure- OH',
                                'DOM- Tea Procure- Blend',
                                'DOM- Speciality- OH',
                                'DOM- Speciality- Blend',
                                'EOU- Tea Procure- OH',
                                'EOU- Tea Procure- Blend',
                                'EOU- Speciality- OH',
                                'EOU- Speciality- Blend']

    required_allocation_CBE = ['Blend Speciality Cost- OH',
                            'Blend Cost Black tea- OH']



    final_filtered_DOM_EOU =  merged_DOM_EOU[merged_DOM_EOU['Allocation'].isin(required_allocation_DOM_EOU)]
    final_filtered_CBE =merged_CBE[merged_CBE['Allocation'].isin(required_allocation_CBE)]


    #####################################################################################################################

    final_filtered_CBE['Departments'] = final_filtered_CBE['Departments'].replace('Tea Instant', 'Specialty')


    master_expenses_df = pd.concat([final_filtered_DOM_EOU, final_filtered_CBE], ignore_index=True)

    exclude_accounts = ['4002600000','4000320000','4000910000']
    master_expenses = master_expenses_df[~master_expenses_df['Account No'].isin(exclude_accounts)]


    mapping = {'Kolkata DOM Branch': 'KOL DOM', 'Kolkata EOU Branch': 'KOL EOU'}

    master_expenses['Branch'] = master_expenses['Branch'].replace(mapping)

    Expenses_raw =master_expenses[['Branch', 'Departments','Account No','Ledger Name','Profit & Loss','Month_Name','Allocation']]

    # Expenses_raw_file= Expenses_raw.drop_duplicates(subset=['Branch', 'Departments', 'Account No', 'Ledger Name', 'Profit & Loss', 'Month_Name'])
    # Expenses_raw_file_reset = Expenses_raw_file.reset_index()

    ##########################################################################################################################

    pivot_table = pd.pivot_table(Expenses_raw, index=['Departments', 'Branch'],columns='Month_Name',values='Profit & Loss'
                                , aggfunc='sum', fill_value=0)
    pivot_table = pivot_table.reindex(columns=month_array)

    pivot_table.fillna(0, inplace=True)

    #########################################################################################################################

    fixed_rent = 510000
    month_columns = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March']
    months_with_data = pivot_table.loc[('Tea Procurement', 'CBE'), month_columns].ne(0)
    for month in months_with_data.index[months_with_data]:
        pivot_table.loc[('Tea Procurement', 'CBE'), month] += fixed_rent


    pivot_table['April-June'] = pivot_table[['April', 'May', 'June']].sum(axis=1)
    pivot_table['July-September'] = pivot_table[['July', 'August', 'September']].sum(axis=1)
    pivot_table['October-December'] = pivot_table[['October', 'November', 'December']].sum(axis=1)
    pivot_table['January-March'] = pivot_table[['January', 'February', 'March']].sum(axis=1)
    pivot_table['YTD'] = pivot_table['April'] + pivot_table['May'] + pivot_table['June'] + pivot_table['July'] + pivot_table['August'] + pivot_table['September'] + pivot_table['October'] + pivot_table['November'] + pivot_table['December'] + pivot_table['January'] + pivot_table['February'] + pivot_table['March']


    department_totals = pivot_table.groupby(level='Departments').sum()


    for department in department_totals.index:
        pivot_table.loc[(department, 'Total'), :] = department_totals.loc[department]


    pivot_table.sort_index(inplace=True)

    pivot_table = pivot_table[['April', 'May', 'June', 'April-June', 'July', 'August', 'September', 'July-September',
                            'October', 'November', 'December', 'October-December', 'January', 'February','March', 'January-March',
                            'YTD']]


    ########################################################################################################################


    pivot_table['April-June'] = pivot_table['April'] + pivot_table['May'] + pivot_table['June']

    pivot_table['July-Sept'] = pivot_table['July'] +pivot_table['August'] + pivot_table['September']

    pivot_table['Oct-Dec'] = pivot_table['October'] + pivot_table['November'] + pivot_table['December']

    pivot_table['Jan-March'] = pivot_table['January'] + pivot_table['February'] + pivot_table['March']

    pivot_table['YTD'] = pivot_table['April'] + pivot_table['May'] + pivot_table['June'] + pivot_table['July'] + pivot_table['August'] + pivot_table['September'] + pivot_table['October'] + pivot_table['November'] + pivot_table['December'] + pivot_table['January'] + pivot_table['February'] + pivot_table['March']

    pivot_table=pivot_table[['April', 'May', 'June', 'April-June', 'July', 'August', 'September', 'July-Sept', 
                            'October', 'November', 'December', 'Oct-Dec', 'January', 'February','March', 'Jan-March',
                            'YTD']]

    ###########################################################################################################################


    ###########################################################################################################################

    # Clean file paths and process files
    cleaned_file_paths = {branch: clean_path(path) for branch, path in file_paths.items()}
    dfs = []

    for branch, path in cleaned_file_paths.items():
        try:
            for file in os.listdir(path):
                if file.endswith('.xlsx') or file.endswith('.xls'):
                    df = pd.read_excel(os.path.join(path, file), header=2)
                    month_name = file.split('_')[1].split('.')[0]  # Extract month name from file name
                    df['Month_Name'] = month_name
                    df['Branch'] = branch
                    dfs.append(df)
        except OSError as e:
            print(f"Error accessing {path}: {e}")

    final_df = pd.concat(dfs, ignore_index=True)


    blend_raw = final_df.drop(['Closed Date', 'Blend No'], axis=1)
    filtered_df =blend_raw[blend_raw['Blend Item Name'].str.contains('Avg Cost', na=False)]


    filtered_df['Blend Item Name'] = filtered_df['Blend Item Name'].str.replace('Avg Cost - ','')

    # filtered_df = filtered_df[~(filtered_df['Blend Item Name'].str.contains('LS', na=False))]
    filtered_df = filtered_df[~(filtered_df['Blend Item Name'].str.contains(r'\bLS\b', na=False))]

    filtered_df['Blend Item Name'] = filtered_df['Blend Item Name'].astype(str)
    


    ####################################################################################################################

    blend_map = blend_map.drop_duplicates(subset=['Chai Item Name', 'Chai Item Category'])
    blend_map['Chai Item Name'] =blend_map['Chai Item Name'].astype(str)

    blend_merge = pd.merge(filtered_df, blend_map, left_on='Blend Item Name', right_on='Chai Item Name', how='left')

    blend_merge = blend_merge[['Branch','Blend Item Name','Chai Item Name', 'Chai Item Category','Physical Qty Kgs.','Month_Name']]
    blend_merge['Blend Item Name'] = blend_merge['Blend Item Name'].astype(str)

    blend_merge = blend_merge[~(blend_merge['Blend Item Name'].str.contains('CHALAI', na=False))]
    # blend_merge.to_excel("D:\\1MJPIL_COIMBATORE\\Blend\\testing4.xlsx")
                                
    departments = []

    for index, row in blend_merge.iterrows():
        category = row['Chai Item Category']

        if category in ['BLENDI', 'BLENDS']:
            departments.append('Specialty')
        elif category in ['BLENDT', 'BLENDB']:
            departments.append('Tea Procurement')
        else:
            departments.append(None)

    blend_merge['Departments'] = departments

    ####################################################################################################################

    blend_table = pd.pivot_table(blend_merge, index=['Departments', 'Branch'], columns='Month_Name', values='Physical Qty Kgs.', aggfunc='sum', fill_value=0)

    blend_table = blend_table.reindex(columns=month_array)

    blend_table.fillna(0, inplace=True)

    blend_table['April-June'] = blend_table['April'] + blend_table['May'] + blend_table['June']

    blend_table['July-Sept'] = blend_table['July'] + blend_table['August'] + blend_table['September']

    blend_table['Oct-Dec'] = blend_table['October'] + blend_table['November'] + blend_table['December']

    blend_table['Jan-March'] = blend_table['January'] + blend_table['February'] + blend_table['March'] 

    blend_table['YTD'] = blend_table['April'] + blend_table['May'] + blend_table['June'] + blend_table['July'] +blend_table['August'] + blend_table['September'] + blend_table['October'] + blend_table['November'] + blend_table['December'] + blend_table['January'] + blend_table['February'] + blend_table['March'] 

    blend_table = blend_table[['April', 'May', 'June', 'April-June', 'July', 'August', 'September', 'July-Sept',
                            'October', 'November', 'December', 'Oct-Dec', 'January', 'February','March', 'Jan-March',
                            'YTD']]

    ########################################################################################################################

    department_branch_totals = blend_table.groupby(level='Departments').sum()


    for department in department_branch_totals.index:
        blend_table.loc[(department, 'Total'), :] = department_branch_totals.loc[department]

    blend_table.sort_index(inplace=True)


    #######################################################################################################################

    rate_per_kg = pd.DataFrame(index=pivot_table.index, columns=pivot_table.columns)

    for index, pivot_row in pivot_table.iterrows():
        departments = index[0]
        branch = index[1]

        if (departments, branch) in blend_table.index:

            blend_row = blend_table.loc[(departments, branch)]

            rate = pivot_row / blend_row
            rate_per_kg.loc[(departments, branch)] = round(rate, 2)
        else:

            rate_per_kg.loc[(departments, branch)] = float('nan')


    rate_last_yr_YTD = rate_last_yr[['Branch','April-March']]
    branch_to_am = rate_last_yr_YTD['April-March'].tolist()

    branch_to_am_series = pd.Series(branch_to_am, index=rate_per_kg.index)
    rate_per_kg['YTD 23-24'] = branch_to_am_series

###############
    error_blend = blend_merge[(blend_merge['Chai Item Category'].isna())]
    error_blend_df = error_blend.drop_duplicates()
    dom_eou_error =error_blend_df[['Branch','Blend Item Name','Chai Item Category']]

    error_exps = blank_check[(blank_check['Allocation'].isna())]
    expns_error =error_exps[['Branch','Departments','Allocation','Account No']]
    error_exps_df = expns_error.drop_duplicates()

    error_cbe = blank_cbe[(blank_cbe['Allocation'].isna())]
    cbe_error =error_cbe[['Branch','Departments','Allocation','Account No']]
    cbe_error_df = cbe_error.drop_duplicates()

###############

    output_file = io.BytesIO()

    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        rate_per_kg.to_excel(writer, sheet_name='Rate_per_Kg', index=True)
        pivot_table.to_excel(writer, sheet_name='Expenses_Volume', startrow=0, startcol=0)
        blend_table.to_excel(writer, sheet_name='Expenses_Volume', startrow=len(pivot_table) + 3, startcol=0)
        blend_merge.to_excel(writer, sheet_name='Blend_File_Raw', index=False)
        Expenses_raw.to_excel(writer, sheet_name='Expenses_File', index=False)
        blank_check.to_excel(writer, sheet_name='DOM_EOU_Raw', index=False)
        blank_cbe.to_excel(writer, sheet_name='CBE_Raw', index=False)
        error_exps_df.to_excel(writer, sheet_name='ERROR_Expenses_Master', index=False)
        cbe_error_df.to_excel(writer, sheet_name='ERROR_CBE_Blend', index=False)
        dom_eou_error.to_excel(writer, sheet_name='ERROR_DOM_EOU_Blend', index=False)

    # Load the workbook from the in-memory buffer
    output_file.seek(0)
    workbook = px.load_workbook(output_file)

    # Define the border style for cells
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Aqua fill color for 'Total' rows
    total_fill = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")

    # Function to autofit column widths
    def autofit_columns(sheet):
        for column in sheet.columns:
            max_length = max(len(str(cell.value)) for cell in column if cell.value is not None)
            adjusted_width = max_length + 5
            column_letter = px.utils.get_column_letter(column[0].column)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Function to format the sheet
    def format_sheet(sheet):
        autofit_columns(sheet)
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical="center", horizontal="center")
                cell.font = Font(size=10)
                # Highlight 'Total' rows
                if cell.value and isinstance(cell.value, str) and 'Total' in cell.value:
                    for cell in row:
                        cell.fill = total_fill
                        cell.font = Font(bold=True)

    # Autofit columns, apply borders, and format sheets
    sheets_to_format = ['Rate_per_Kg', 'Blend_File_Raw', 'Expenses_Volume', 'Expenses_File']
    for sheet_name in sheets_to_format:
        sheet = workbook[sheet_name]
        format_sheet(sheet)

    # Expenses Volume sheet - Comma Style and Thousands Separator
    expenses_volume_sheet = workbook['Expenses_Volume']
    for row in expenses_volume_sheet.iter_rows(min_row=2, max_row=expenses_volume_sheet.max_row, min_col=1):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = '#,##0'  # Apply comma style formatting for numbers

    # Bold Headers for Rate_per_Kg and Expenses_Volume
    rate_per_kg_sheet = workbook['Rate_per_Kg']
    rate_per_kg_sheet['A1'].font = Font(size=11, bold=True)
    rate_per_kg_sheet['A1'].value = 'Rate/KG'

    # Expenses Volume sheet header formatting
    expenses_volume_sheet['A1'].value = 'EXPENSES'
    expenses_volume_sheet['A1'].font = Font(size=11, bold=True)

    # Setting specific headers for sections in the Expenses_Volume sheet
    header_cell_volume = expenses_volume_sheet.cell(row=len(pivot_table) + 4, column=1)
    header_cell_volume.value = 'VOLUME'
    header_cell_volume.font = Font(size=11, bold=True)

    # Set row height for Expenses Volume sheet
    for row in expenses_volume_sheet.iter_rows(min_row=1, max_row=expenses_volume_sheet.max_row):
        expenses_volume_sheet.row_dimensions[row[0].row].height = 20

    # Save the modified workbook to the in-memory buffer
    output_file.seek(0)  # Reset the buffer before saving
    new_output_file = io.BytesIO()  # Create a fresh buffer for the final output
    workbook.save(new_output_file)
    new_output_file.seek(0)

    # Store the file in the Flask app config
    current_app.config['EXCEL_FILE'] = {
        'name': 'BLEND Overhead Report.xlsx',
        'data': new_output_file.getvalue()
    }

    # Redirect to the processing page
    return redirect(url_for('processing'))

@blendovr.route('/download')
def download_report():
    # Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])

@blendovr.route('/blendovr-master', methods=['GET'])
def blendovr_master():
    # Path to the default lastyear sales file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\Blend Mapping Master.xlsx'
    return send_file(default_allocation_path, as_attachment=True)

@blendovr.route('/blendovr-lastyr', methods=['GET'])
def blendovr_lastyr():
    # Path to the default lastyear sales file
    default_lastyr_path = r'D:\2MJIPL_COIMBATORE\Last Year Blend Overhead.xlsx'
    return send_file(default_lastyr_path, as_attachment=True)

from io import BytesIO
# Regular expression to extract the month name from the filename
month_pattern = re.compile(r'(April|May|June|July|August|September|October|November|December|January|February|March)', re.IGNORECASE)

# Financial year month order (April to March)
financial_month_order = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March']


@blendovr.route('/process_filesname', methods=['GET'])
def process_files():
    # Retrieve subfolders
    subfolders = [f.path for f in os.scandir(UPLOAD_FOLDER) if f.is_dir()]

    # Dictionary to hold subfolder names and their corresponding files
    subfolder_files = {}

    for subfolder in subfolders:
        # Get the subfolder name (without the full path)
        subfolder_name = os.path.basename(subfolder)
        try:
            # List all files in the subfolder
            files = os.listdir(subfolder)
            
            # Filter and match files by month names
            matched_files = []
            for file in files:
                match = month_pattern.search(file)
                if match:
                    matched_files.append(match.group(0).capitalize())  # Capitalize month name
            
            # Sort the matched files by financial year month order
            matched_files_sorted = sorted(matched_files, key=lambda x: financial_month_order.index(x))
            
            # Store the sorted files in the dictionary
            subfolder_files[subfolder_name] = matched_files_sorted

        except Exception as e:
            return f"An error occurred while processing files in '{subfolder_name}': {str(e)}", 500

    # If no subfolders found, return an error
    if not subfolder_files:
        return "No subfolders with matching files found.", 404

    # Create a DataFrame with subfolders as columns and months as rows
    max_rows = max(len(files) for files in subfolder_files.values())  # Get max number of rows (months)
    df = pd.DataFrame({subfolder: pd.Series(files) for subfolder, files in subfolder_files.items()})

    # Fill missing rows with empty values
    df = df.fillna('')

    # Write DataFrame to an Excel file in memory
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Available Months')
        output.seek(0)

        # Return the Excel file as an attachment for download
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='Available_Months.xlsx')

    except Exception as e:
        return f"An error occurred while generating the Excel file: {str(e)}", 500

    
if __name__ == '__main__':
    blendovr.run(debug=True)

