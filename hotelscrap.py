from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import requests
from bs4 import BeautifulSoup
import pandas as pd
import os

googlescp = Blueprint('googlescp', __name__)

# Function to apply borders to the worksheet
def apply_borders(worksheet):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

# Function to auto-fit columns and apply formatting
def auto_fit_columns(worksheet):
    for col in worksheet.columns:
        max_length = max(len(str(worksheet.cell(row=1, column=col[0].column).value)),
                         max(len(str(cell.value)) for cell in col if cell.value is not None))
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Your scraping function
def scrape_data(url):
    hotels = []
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36"
    }
    response = requests.get(url, headers=headers)
    if response.status_code == 200:       
        soup = BeautifulSoup(response.content, 'html.parser')
        hotel_cards = soup.find_all('div', class_='Zvwhrc')
        
        for card in hotel_cards:           
            single_page_url = "https://www.google.com"            
            anchor_tag = card.find('a', class_='OxGZuc W8vlAc lRagtb')
            href_link = anchor_tag['href'] if anchor_tag else ''
            sungle_url = f"{single_page_url}{href_link}"

            hotel_name = card.find('div', class_='QT7m7').text.strip()
            price_elem = card.find('span', class_='qQOQpe prxS3d')
            hotel_price = price_elem.text.strip() if price_elem else 'N/A'
            rating_elem = card.find('span', class_='UqrZme sSHqwe ogfYpf')
            hotel_rating = rating_elem.text.strip() if rating_elem else 'N/A'
           
            single_page_response = requests.get(sungle_url)
            single_page_soup = BeautifulSoup(single_page_response.content, 'html.parser')

            address_elem = single_page_soup.find('div', class_='K4nuhf')
            if address_elem:
                spans = address_elem.find_all('span', class_='CFH2De')
                if len(spans) >= 2:
                    address = spans[0].text.strip()
                    mobile = spans[2].text.strip() if len(spans) > 2 else 'N/A'
                else:
                    address = 'N/A'
                    mobile = 'N/A'
            else:
                address = 'N/A'
                mobile = 'N/A'

            hotels.append({
                'Name': hotel_name,
                'Rating': hotel_rating,
                'Price': hotel_price,
                'Address': address,
                'Mobile. No': mobile
            })
    return hotels


@googlescp.route('/')
def index():
    return render_template('test.html')

@googlescp.route('/scrapegoogle', methods=['POST'])
def scrape():
    urls = request.form.get('urls')
    if urls:
        # Split URLs by comma (in case multiple URLs are provided)
        url_list = [url.strip() for url in urls.split(',')]
        all_hotels_data = []

        # Scrape each URL
        for url in url_list:
            hotels_data = scrape_data(url)
            all_hotels_data.extend(hotels_data)

        # Convert the collected data to a DataFrame
        df = pd.DataFrame(all_hotels_data)
        df = df.drop_duplicates()

        # Save to Excel file
        output_path = io.BytesIO()
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.round().to_excel(writer, sheet_name='Data', index=True)

            workbook = writer.book
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                auto_fit_columns(worksheet)
                apply_borders(worksheet)

        output_path.seek(0)  # Reset file pointer to the beginning of the file

        current_app.config['EXCEL_FILE'] = {
            'name': 'From Google Sites.xlsx',
            'data': output_path.getvalue()}


        # Redirect to the processing page
        return redirect(url_for('processing'))

@googlescp.route('/processing')
def processing():
    return render_template('processing.html')

@googlescp.route('/download')
def download_report():
# Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])

if __name__ == "__main__":
    googlescp.run(debug=True)
