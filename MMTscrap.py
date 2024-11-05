from flask import Flask, Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import requests

import time
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from bs4 import BeautifulSoup

mmtscp = Blueprint('mmtscp', __name__)

# Function to apply borders to the worksheet
def apply_borders(worksheet):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

# Function to auto-fit columns and apply formatting
def auto_fit_columns(worksheet):
    for col in worksheet.columns:
        max_length = max(len(str(worksheet.cell(row=1, column=col[0].column).value)),
                         max(len(str(cell.value)) for cell in col if cell.value is not None))
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Route for homepage
@mmtscp.route('/')
def index():
    return render_template('test.html')

# Route to handle scraping and file download
@mmtscp.route('/scrapemmt', methods=['POST'])
def scrape():
    if request.method == 'POST':
        # Get URL from the form input
        url = request.form['url']
        
        # Set up Chrome WebDriver
        chrome_options = webdriver.ChromeOptions()
        chrome_options.add_argument("--disable-notifications")  # Disable notifications (to prevent pop-ups)
        driver = webdriver.Chrome(options=chrome_options)
        driver.set_window_size(1024, 768)
        
        # Navigate to the tracking page (URL from user)
        driver.get(url)

        # Scroll and Extract Data
        scroll_script = "window.scrollTo(0, document.body.scrollHeight);"
        while True:
            try:
                driver.execute_script(scroll_script)
                WebDriverWait(driver, 50).until(
                    EC.presence_of_element_located((By.XPATH, "//span[text()='Load More']"))
                )
            except Exception as e:
                break

        # Extract data from the final page source
        page_source = driver.page_source
        soup = BeautifulSoup(page_source, "html.parser")

        names = []
        prices = []
        areas = []

        hotel_names = soup.find_all("span", class_="wordBreak appendRight10")
        for hotel in hotel_names:
            hotel_name = hotel.text.strip()
            if not hotel_name:
                hotel_name = "0"  # Replace empty name with "0"
            names.append(hotel_name)

        hotel_prices = soup.find_all("p", class_="priceText latoBlack font22 blackText appendBottom5")
        for hotelprice in hotel_prices:
            hotel_price = hotelprice.text.strip()
            if not hotel_price:
                hotel_price = "0"  # Replace empty name with "0"
            else:
                hotel_price = ''.join(filter(str.isdigit, hotel_price))
            prices.append(hotel_price)

        hotel_area_div = soup.find_all("div", class_="pc__html")
        for hotelarea in hotel_area_div:
            hotel_area = hotelarea.find("span", class_="blueText")
            if hotel_area is not None:
                hotelarea = hotel_area.text.strip()
                areas.append(hotelarea)

        # Create a DataFrame
        df = pd.DataFrame({"Hotel Name": names, "Hotel price": prices, "Area": areas})

        # Save to Excel file
        output_path = io.BytesIO()
        with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
            df.round().to_excel(writer, sheet_name='Data', index=True)

            workbook = writer.book
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                auto_fit_columns(worksheet)
                apply_borders(worksheet)

        output_path.seek(0)  # Reset file pointer to the beginning of the file

        current_app.config['EXCEL_FILE'] = {
            'name': 'From MMT Scrape.xlsx',
            'data': output_path.getvalue()}
        
        driver.quit()  # Close the browser


        # Redirect to the processing page
        return redirect(url_for('processing'))


@mmtscp.route('/processing')
def processing():
    return render_template('processing.html')

@mmtscp.route('/download')
def download_report():
# Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])

if __name__ == "__main__":
    mmtscp.run(debug=True)

