from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
import re
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import xlsxwriter
from tempfile import NamedTemporaryFile
import pandas as pd
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl import load_workbook

cyly = Blueprint('cyly', __name__)

@cyly.route('/')
def index():
    return render_template('test.html')

@cyly.route('/cyly-upload', methods=['POST'])
def upload_files():
    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\ITEM DESCRIPTION WITH CATEGORY (map).xlsx'
    # Check if an allocation file was provided by the user
    allocation_file = request.files.get('allocation')
    if allocation_file:
        # If the user provided an allocation file, delete the existing default file
        if os.path.exists(default_allocation_path):
            os.remove(default_allocation_path)  # Delete the existing default allocation file
        allocation_file.save(default_allocation_path)  # Save the new allocation file
        category_allocation = pd.read_excel(allocation_file)
    else:
        # If no allocation file is provided, use the default allocation file
        category_allocation = pd.read_excel(default_allocation_path)


    # Path to the default last year sales file
    default_lastyr_path = r'D:\2MJIPL_COIMBATORE\Sales Register April23 to March24.xlsx'
    # Check if a last year sales file was provided by the user
    lastyr_file = request.files.get('salesly')
    if lastyr_file:
        # If the user provided a last year sales, delete the existing default file
        if os.path.exists(default_lastyr_path):
            os.remove(default_lastyr_path)  # Delete the existing default_lastyr_path
        lastyr_file.save(default_lastyr_path)  # Save the new last year sales file
        sales_report_file23_24 = pd.read_excel(lastyr_file)
    else:
        # If no allocation file is provided, use the default allocation file
        sales_report_file23_24 = pd.read_excel(default_lastyr_path)        
    

    # Read the uploaded Excel files

    sales_cy = request.files['salescy']

    sales_report_file24_25 = pd.read_excel(sales_cy)

  

    category_allocation['Allocation'] = category_allocation['Allocation'].str.upper()

    sales_report_file23_24['Phyto Date'] = pd.to_datetime(sales_report_file23_24['Phyto Date'])
    sales_report_file23_24['Month'] = sales_report_file23_24['Phyto Date'].dt.strftime('%B')
    sales_report_file24_25['Phyto Date'] = pd.to_datetime(sales_report_file24_25['Phyto Date'])
    sales_report_file24_25['Month'] = sales_report_file24_25['Phyto Date'].dt.strftime('%B')

    ######################################################################################################################################################

    current_year_months = set(sales_report_file24_25['Month'].unique())
    previous_year_months = set(sales_report_file23_24['Month'].unique())

    common_months = current_year_months.intersection(previous_year_months)

    previous_year_data23_24 = sales_report_file23_24[sales_report_file23_24['Month'].isin(common_months)]
    # previous_year_data23_24['Month'].unique()


    #####################################################################################################################################################

    # GL_Description = [
    #     '410001 - Sales Tea Export',
    #     '410002 - Sales Tea Domestic',
    #     '410029 - Export Sales Schedule 3']
    GL_Description = [
        '3000040000 - Sales Tea Export',
        '3000100000 - Sales Tea Domestic',
        '3000070000 - Export Sales Schedule 3']


    sales_report_file24_25['G/L Ac-Description'] =sales_report_file24_25['G/L Ac-Description'].str.strip()

    GL_Description_stripped = [desc.strip() for desc in GL_Description]

    sales_report_GL_desc_filter = sales_report_file24_25[sales_report_file24_25['G/L Ac-Description'].isin(GL_Description_stripped)]

    sales_report_GL_desc_filter['Year'] = 'CY'

    sales_report_GL_desc_filter.loc[sales_report_GL_desc_filter['Department'].str.contains(r'^(RED|Red)\b'), 'Department'] = 'Total Red Sales'

    sales_report_GL_desc_filter.loc[sales_report_GL_desc_filter['Department'].str.contains(r'^(YELLOW|Yellow)\b'), 'Department'] = 'Total Yellow Sales'


    def extract_sales(text):
        match = re.search(r'(Total Red Sales|Total Yellow Sales)', text)
        if match:
            return match.group(0)
        else:
            return text

    sales_report_GL_desc_filter['Department'] = sales_report_GL_desc_filter['Department'].apply(extract_sales)

    def get_segment_department(row):
        if 'Horeca Sales' in row['Department'] or 'Online Sales' in row['Department'] or 'Private Lable Sales' in row['Department'] or 'TEA ME Sales' in row['Department'] or 'Total Red Sales' in row['Department'] or 'Total Yellow Sales' in row['Department']:
            return 'Domestic'
        elif 'Export Africa Sales' in row['Department'] or 'Export AUS/EUR Sales' in row['Department'] or 'US sales' in row['Department'] or 'Other Export Sales' in row['Department']:
            return 'Export'
        elif 'Russia Trade' in row['Department'] or 'Abidjan Trade' in row['Department'] or 'Ghana Trade' in row['Department']:
            return 'Merchant Trade Sales'
        else:
            return 'Other'

    sales_report_GL_desc_filter['Segment Departments'] = sales_report_GL_desc_filter.apply(get_segment_department, axis=1)
    sales_report_customer_filter = sales_report_GL_desc_filter[~sales_report_GL_desc_filter['Customer Name'].str.lower().str.startswith('madhu jayanti international'.lower())]

    category_filter = category_allocation.drop_duplicates(subset=['Item Description'])
    CY_sales = pd.merge(sales_report_customer_filter, category_filter, on='Item Description', how='left')
    CY_sales = CY_sales[CY_sales['Segment Departments'] != 'Other']


    CY_sales['Customer Country'] = CY_sales['Customer Country'].fillna(' -- BLANK --')


    #some changes in CY for match with LY
    # Strip whitespace and convert to uppercase
    CY_sales['Customer Name'] = CY_sales['Customer Name'].str.strip().str.upper()
    CY_sales['Customer Country'] = CY_sales['Customer Country'].str.strip().str.upper()
    CY_sales['Customer Country'].replace('UNITED STATES', 'USA', inplace=True)
    CY_sales['Department'].replace('UNITED STATES Sales','USA Sales', inplace = True)


    ####################################################################################################################################################

    GL_desc_include =['410001 - Sales Tea Export',
        '410002 - Sales Tea Domestic',
        '410029 - Export Sales Schedule 3']

    previous_year_data23_24['G/L Ac-Description'] =previous_year_data23_24['G/L Ac-Description'].str.strip()
    GL_des_stripped = [desc.strip() for desc in GL_desc_include]
    GL_desc_filter =previous_year_data23_24[previous_year_data23_24['G/L Ac-Description'].isin(GL_des_stripped)]

    GL_desc_filter['Year'] = 'LY'

    GL_desc_filter.loc[GL_desc_filter['Department'].str.contains(r'^(RED|Red)\b'), 'Department'] = 'Total Red Sales'

    GL_desc_filter.loc[GL_desc_filter['Department'].str.contains(r'^(YELLOW|Yellow)\b'), 'Department'] = 'Total Yellow Sales'

    def extract_sales_dep(text):
        matching = re.search(r'(Total Red Sales|Total Yellow Sales)', text)
        if matching:
            return matching.group(0)
        else:
            return text

    GL_desc_filter['Department'] = GL_desc_filter['Department'].apply(extract_sales_dep)

    def get_segment_department_value(row):
        if 'Horeca Sales' in row['Department'] or 'Online Sales' in row['Department'] or 'Private Lable Sales' in row['Department'] or 'TEA ME Sales' in row['Department'] or 'Total Red Sales' in row['Department'] or 'Total Yellow Sales' in row['Department']:
            return 'Domestic'
        elif 'Export Africa Sales' in row['Department'] or 'Export AUS/EUR Sales' in row['Department'] or 'US sales' in row['Department'] or 'Other Export Sales' in row['Department']:
            return 'Export'
        elif 'Russia Trade' in row['Department'] or 'Abidjan Trade' in row['Department'] or 'Ghana Trade' in row['Department']:
            return 'Merchant Trade Sales'
        else:
            return 'Other'
        
    GL_desc_filter['Segment Departments'] = GL_desc_filter.apply(get_segment_department_value, axis=1)
    customer_name_filter = GL_desc_filter[~GL_desc_filter['Customer Name'].str.lower().str.startswith('madhu jayanti international'.lower())]

    category_filter = category_allocation.drop_duplicates(subset=['Item Description'])
    LY_sales = pd.merge(customer_name_filter , category_filter, on='Item Description', how='left')
    LY_sales = LY_sales[LY_sales['Segment Departments'] != 'Other']

    # Strip whitespace and convert to uppercase
    LY_sales['Customer Name'] = LY_sales['Customer Name'].str.strip().str.upper()
    LY_sales['Customer Country'] = LY_sales['Customer Country'].str.strip().str.upper()

    #######################################################################################################################################################

    marge_raw = pd.concat([CY_sales, LY_sales], axis=0, ignore_index=True)
    marge_raw = marge_raw.rename(columns={'Allocation': 'Category'})
    marge_raw['Date'] = marge_raw['Phyto Date'].dt.date

    raw_file = marge_raw[['Item Description','G/L Ac-Description', 'Department','Segment Departments', 'Category','Customer Name','Customer Country','Date', 'Year','Sale Amount Actual','Total Net Wt',]]

    #######################################################################################################################################################
    #Definer Amount
    rs_df = marge_raw[['Segment Departments', 'Department', 'Category', 'Year', 'Sale Amount Actual']]
    df_total = rs_df.groupby(["Segment Departments", "Year"])['Sale Amount Actual'].sum().reset_index()

    df_total_List = []

    for seg_dpt in df_total["Segment Departments"].unique():
        temp_df = df_total[df_total["Segment Departments"] == seg_dpt].copy()
        temp_df['Department'] =f"Year Wise {seg_dpt} Total"
        temp_df['Category'] = ' '
        df_total_List.append(temp_df)


        grand_total_row = pd.DataFrame({
            'Segment Departments': [seg_dpt],
            'Department': [f"Year Wise {seg_dpt} Total"],
            'Category': [' '],
    #         'Year': [' '],
            'Sale Amount Actual': [temp_df['Sale Amount Actual'].sum()]
        })
        df_total_List.append(grand_total_row)

    rs_with_dpt_total = pd.concat([rs_df] + df_total_List)

    category_total = rs_df.groupby(['Segment Departments', 'Department', 'Year'])['Sale Amount Actual'].sum().reset_index()
    category_total['Category'] = 'TOTAL'
    rs_with_Category_total = pd.concat([rs_with_dpt_total, category_total])

    grand_total = rs_df.groupby(['Year'])['Sale Amount Actual'].sum().reset_index()
    grand_total['Segment Departments'] = 'Year Wise GRAND TOTAL'
    grand_total['Department'] = ' '
    grand_total['Category'] = ' '

    proper_table = pd.concat([rs_with_Category_total, grand_total])
    proper_table.sort_values(by=('Segment Departments'), ascending=False, inplace=True)

    INR_df = pd.pivot_table(proper_table, index=['Segment Departments', 'Department', 'Category'],
                            columns='Year', values='Sale Amount Actual',
                            aggfunc='sum', fill_value=0)
    INR_df['Deviation% (INR)'] = ((INR_df['CY'] - INR_df['LY']) / INR_df['LY']) * 100
    INR_df['Deviation% (INR)'] = INR_df['Deviation% (INR)'].replace([np.inf, -np.inf], 100)
    INR_df.rename(columns={'CY': 'CY (INR)', 'LY': 'LY (INR)'}, inplace=True)


    # Define kg_df
    kg_df = marge_raw[['Segment Departments', 'Department', 'Category', 'Year', 'Total Net Wt']]


    df_quantity = kg_df.groupby(["Segment Departments", "Year"])['Total Net Wt'].sum().reset_index()

    df_total_List = []

    for seg_dpt in df_quantity["Segment Departments"].unique():
        temp_df = df_quantity[df_quantity["Segment Departments"] == seg_dpt].copy()
        
        temp_df['Department'] = f"Year Wise {seg_dpt} Total"
        temp_df['Category'] = ' '
        df_total_List.append(temp_df)
        
        grand_total_row = pd.DataFrame({
            'Segment Departments': [seg_dpt],
            'Department': [f"Year Wise {seg_dpt} Total"],
            'Category': [' '],
    #         'Year': [' '],
            'Total Net Wt': [temp_df['Total Net Wt'].sum()]
        })
        df_total_List.append(grand_total_row)

    kg_with_dpt_total = pd.concat([kg_df] + df_total_List)


    category_total = kg_df.groupby(['Segment Departments', 'Department', 'Year'])['Total Net Wt'].sum().reset_index()
    category_total['Category'] = 'TOTAL'
    kg_with_Category_total = pd.concat([kg_with_dpt_total, category_total])


    grand_total_kg = kg_df.groupby(['Year'])['Total Net Wt'].sum().reset_index()
    grand_total_kg['Segment Departments'] = 'Year Wise GRAND TOTAL'
    grand_total_kg['Department'] = ' '
    grand_total_kg['Category'] = ' '

    proper_table_kg = pd.concat([kg_with_Category_total, grand_total_kg])

    proper_table_kg.sort_values(by=('Segment Departments'), ascending=False, inplace=True)

    KG_df = pd.pivot_table(proper_table_kg, index=['Segment Departments', 'Department', 'Category'],
                        columns='Year', values='Total Net Wt', aggfunc='sum', fill_value=0)


    KG_df['Deviation% (KG)'] = ((KG_df['CY'] - KG_df['LY']) / KG_df['LY']) * 100
    KG_df['Deviation% (KG)'] = KG_df['Deviation% (KG)'].replace([np.inf, -np.inf], 100)

    KG_df.rename(columns={'CY': 'CY (KG)', 'LY': 'LY (KG)'}, inplace=True)


    merged_table = INR_df.merge(KG_df, left_index=True, right_index=True)
    merged_table['CY (Rate/KG)'] = merged_table['CY (INR)']/merged_table['CY (KG)']
    merged_table['LY (Rate/KG)'] = merged_table['LY (INR)']/merged_table['LY (KG)']
    merged_table['LY (Rate/KG)'] = merged_table['LY (Rate/KG)'].fillna(0)
    merged_table['Deviation% (Rate/KG)'] = ((merged_table['CY (Rate/KG)']-merged_table['LY (Rate/KG)'])/merged_table['LY (Rate/KG)'])*100
    merged_table['Deviation% (Rate/KG)'] = merged_table['Deviation% (Rate/KG)'].replace([np.inf, -np.inf], 100)

    ######################################################################################################################################################


    def format_number(x):
        if isinstance(x, (int, float)):
            if np.isinf(x) or np.isnan(x):
                return np.nan
            return int(x)
        else:
            return x

    # Function to format percentages
    def format_percentage(x):
        if isinstance(x, (int, float)):
            return f"{x:.2f}%"
        elif x == float('inf'):
            return 'inf%'
        else:
            return x

    merged_table[['CY (INR)', 'LY (INR)', 'CY (KG)', 'LY (KG)', 'CY (Rate/KG)', 'LY (Rate/KG)']] = \
        merged_table[['CY (INR)', 'LY (INR)', 'CY (KG)', 'LY (KG)', 'CY (Rate/KG)', 'LY (Rate/KG)']].applymap(format_number)

    merged_table[['Deviation% (INR)', 'Deviation% (KG)', 'Deviation% (Rate/KG)']] = \
        merged_table[['Deviation% (INR)', 'Deviation% (KG)', 'Deviation% (Rate/KG)']].applymap(format_percentage)
    merged_table


    ##################################################################################################################################################

    #FOR CY#
    export_sales_CY = CY_sales[(CY_sales['Segment Departments'] == 'Export') & (CY_sales['Customer Country'] != 'USA')]

    export_group = export_sales_CY.groupby(['Department', 'Customer Name', 'Customer Country']).agg({
        'Sale Amount Actual': 'sum',
        'Total Net Wt': 'sum'
    }).reset_index()

    export_group.sort_values(by=['Sale Amount Actual', 'Total Net Wt'], ascending=False, inplace=True)

    department_totals = export_group.groupby('Department').sum().reset_index()
    department_totals['Customer Name'] = 'TOTAL'
    department_totals['Customer Country'] = ' '

    export_group_with_totals = []

    for department in export_group['Department'].unique():
        temp_df = export_group[export_group['Department'] == department].copy()
        export_group_with_totals.append(temp_df)
        
        temp_total_row = department_totals[department_totals['Department'] == department]
        export_group_with_totals.append(temp_total_row)

    export_group_with_totals = pd.concat(export_group_with_totals)

    export_group_with_totals.rename(columns={'Customer Country': ' Country ', 'Sale Amount Actual': 'CY (INR)', 'Total Net Wt': 'CY (KG)'}, inplace=True)

    # FOR LY#
    export_sales_LY =LY_sales[(LY_sales['Segment Departments'] == 'Export') & (LY_sales['Customer Country'] != 'USA')]


    exportdata_group = export_sales_LY.groupby(['Department', 'Customer Name','Customer Country']).agg({
        'Sale Amount Actual': 'sum',
        'Total Net Wt': 'sum'
    }).reset_index()  
    

    exportdata_group.sort_values(by=(['Sale Amount Actual', 'Total Net Wt']), ascending=False, inplace=True)

    department_group = exportdata_group.groupby('Department').agg({
        'Sale Amount Actual': 'sum',
        'Total Net Wt': 'sum',
        'Customer Name': lambda x: 'TOTAL',
        'Customer Country': lambda x: ''
    }).reset_index()

    export_group_with_total_data = []

    for department in exportdata_group['Department'].unique():
        temp_data =exportdata_group[exportdata_group['Department'] == department].copy()

        export_group_with_total_data .append(temp_data)
        
        temp_total_row_data =department_group[department_group['Department'] == department]
        export_group_with_total_data .append(temp_total_row_data)

    export_group_with_total_data = pd.concat(export_group_with_total_data)

    export_group_with_total_data.rename(columns={'Customer Country': ' Country ', 'Sale Amount Actual': 'LY (INR)','Total Net Wt': 'LY (KG)'}, inplace=True)


    #####################################################################################################################################################


    export_sales_CY = CY_sales[(CY_sales['Segment Departments'] == 'Export') & (CY_sales['Customer Country'] == 'USA')]


    export_group_CY = export_sales_CY.groupby(['Department', 'Customer Name', 'Customer Country']).agg({
        'Sale Amount Actual': 'sum',
        'Total Net Wt': 'sum'
    }).reset_index()

    # export_group_CY['Department'] = export_group_CY.apply(lambda row: 'USA Sales' if row['Customer Country'] == 'USA' else f"{row['Customer Country']} Sales", axis=1)


    export_sales_LY = LY_sales[(LY_sales['Segment Departments'] == 'Export') & (LY_sales['Customer Country'] == 'USA')]
    export_group_LY = export_sales_LY.groupby(['Department', 'Customer Name', 'Customer Country']).agg({
        'Sale Amount Actual': 'sum',
        'Total Net Wt': 'sum'
    }).reset_index()



    usa_df = pd.merge(export_group_CY, export_group_LY, on = ['Department','Customer Name' , 'Customer Country'],how='outer')

    usa_df.rename(columns={'Customer Country': ' Country ', 'Sale Amount Actual_x': 'CY (INR)',
                            'Sale Amount Actual_y': 'LY (INR)','Total Net Wt_x': 'CY (KG)','Total Net Wt_y': 'LY (KG)'}, inplace=True)
    usa_df = usa_df.fillna(0)
    usa_df['Department'] = usa_df.apply(lambda row: 'USA Sales' if row[' Country '] == 'USA' else f"{row[' Country ']} Sales", axis=1)

    usa_df['Customer Name'] = usa_df['Customer Name'].str.replace(r'\s+', ' ', regex=True)
    usa_df['Customer Name'] = usa_df['Customer Name'].str.strip()

    total_cy_inr = usa_df['CY (INR)'].sum()
    total_cy_kg = usa_df['CY (KG)'].sum()
    total_ly_inr = usa_df['LY (INR)'].sum()
    total_ly_kg = usa_df['LY (KG)'].sum()

    usa_df = usa_df.groupby(['Department', 'Customer Name',' Country ']).agg({
        'CY (INR)': 'sum',
        'CY (KG)': 'sum',
        'LY (INR)': 'sum',
        'LY (KG)': 'sum'
    }).reset_index()


    total_row = pd.DataFrame({
        'Department': [' '],
        'Customer Name': ['TOTAL'],
        ' Country ': ['  '],
        'CY (INR)': [total_cy_inr],
        'CY (KG)': [total_cy_kg],
        'LY (INR)': [total_ly_inr],
        'LY (KG)': [total_ly_kg]
    })

    usa_data = pd.concat([usa_df, total_row], ignore_index=True)
    usa_data =usa_data.fillna(0)


    #####################################################################################################################################################

    merged_df = pd.merge(export_group_with_totals, export_group_with_total_data, on=['Department','Customer Name', ' Country '], how='outer')

    result_dfs = []
    for department in merged_df['Department'].unique():
        department_df = merged_df[merged_df['Department'] == department]
        customers_df = department_df[department_df['Customer Name'] != 'TOTAL']
        total_df = department_df[department_df['Customer Name'] == 'TOTAL']
        
        if len(total_df) > 1:
            cy_inr = total_df['CY (INR)'].sum()
            cy_kg = total_df['CY (KG)'].sum()
            ly_inr = total_df['LY (INR)'].sum()
            ly_kg = total_df['LY (KG)'].sum()

            new_total_row = pd.DataFrame({
                'Department': [department],
                'Customer Name': ['TOTAL'],
                ' Country ': [''],
                'CY (INR)': [cy_inr],
                'CY (KG)': [cy_kg],
                'LY (INR)': [ly_inr],
                'LY (KG)': [ly_kg]})
            customers_df = pd.concat([customers_df, new_total_row], ignore_index=True)
        result_dfs.append(customers_df)
    result_df = pd.concat(result_dfs, ignore_index=True)
    final_df = pd.concat([result_df, usa_data], ignore_index=True)

    grand_total =final_df[final_df['Customer Name'] == 'TOTAL'].sum()
    grand_total['Department'] = ' GRAND  TOTAL '
    grand_total['Customer Name'] = ' '
    grand_total = pd.DataFrame(grand_total).transpose()

    export_data = pd.concat([final_df,grand_total],ignore_index=True)
    export_data.fillna(0)
    def calculate_deviation_inr(row):
        if not np.isnan(row['CY (INR)']) and not np.isnan(row['LY (INR)']) and row['LY (INR)'] != 0:
            return ((row['CY (INR)'] - row['LY (INR)']) / row['LY (INR)']) * 100
        elif np.isnan(row['CY (INR)']) and not np.isnan(row['LY (INR)']):
            return 0
        elif not np.isnan(row['CY (INR)']) and np.isnan(row['LY (INR)']):
            return 100
        else:
            return np.nan

    def calculate_deviation_kg(row):
        if not np.isnan(row['CY (KG)']) and not np.isnan(row['LY (KG)']) and row['LY (KG)'] != 0:
            return ((row['CY (KG)'] - row['LY (KG)']) / row['LY (KG)']) * 100
        elif np.isnan(row['CY (KG)']) and not np.isnan(row['LY (KG)']):
            return 0
        elif not np.isnan(row['CY (KG)']) and np.isnan(row['LY (KG)']):
            return 100
        else:
            return np.nan

    export_data['Deviation% (INR)'] = export_data.apply(calculate_deviation_inr, axis=1)
    export_data['Deviation% (KG)'] = export_data.apply(calculate_deviation_kg, axis=1)

    export_data = export_data.fillna(0)

    arrange_columns = ['Department','Customer Name',' Country ','CY (INR)','LY (INR)','Deviation% (INR)','CY (KG)','LY (KG)','Deviation% (KG)']
    export_sheet = export_data[arrange_columns]

    ##################################################################################################################################################

    # Function to format numbers
    def format_number(x):
        if isinstance(x, (int, float)):
            if np.isinf(x) or np.isnan(x):
                return np.nan
            return int(x)
        else:
            return x

    # Function to format percentages
    def format_percentage_add(x):
        if isinstance(x, (int, float)):
            if np.isinf(x):
                return 'inf%'
            return f"{x:.2f}%"
        else:
            return x

    export_sheet[['CY (INR)', 'LY (INR)', 'CY (KG)', 'LY (KG)']] = \
        export_sheet[['CY (INR)', 'LY (INR)', 'CY (KG)', 'LY (KG)']].applymap(format_number)

    export_sheet[['Deviation% (INR)', 'Deviation% (KG)']] = \
        export_sheet[['Deviation% (INR)', 'Deviation% (KG)']].applymap(format_percentage_add)
    

    error_df = raw_file[['Item Description','Category']]
    error_file = error_df[(error_df['Category'].isna())]
    error_file = error_file.drop_duplicates()

    output_file = io.BytesIO()

    # Create a writer object using the openpyxl engine
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        export_sheet.to_excel(writer, sheet_name='Export_Country_wise_Sales', index=False)
        merged_table.to_excel(writer, sheet_name='CY_LY_sales_comparison', index=True)
        raw_file.to_excel(writer, sheet_name='Raw_File_Total_Sales', index=False)
        error_file.to_excel(writer, sheet_name='ERROR_Rows', index=False)

    # Make sure the writer is closed before reloading the workbook
    output_file.seek(0)

    # Load the workbook to apply formatting
    wb = load_workbook(output_file)

    sheets_to_format = ['Export_Country_wise_Sales', 'CY_LY_sales_comparison', 'Raw_File_Total_Sales']

    # Define the styles for formatting
    bold_font = Font(bold=True)
    aqua_fill = PatternFill(start_color="D6ECFC", end_color="D6ECFC", fill_type="solid")
    orange_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") 

    border_style = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Comma style for numbers with thousands separator
    def apply_comma_format(sheet):
        for row in sheet.iter_rows(min_row=2, min_col=1):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.border = border_style

    # Auto-fit column width
    def autofit_columns(sheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Apply formatting for 'Total' or 'TOTAL' rows
    def format_total_rows(sheet):
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if any(cell.value in ['Total', 'TOTAL'] for cell in row):
                for cell in row:
                    cell.font = bold_font
                    cell.fill = aqua_fill
                    cell.border = border_style

    # Apply formatting for 'Year Wise GRAND TOTAL' and 'GRAND TOTAL' rows
    def format_grand_total_rows(sheet):
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            if any(cell.value in ['Year Wise GRAND TOTAL', ' GRAND TOTAL'] for cell in row):
                for cell in row:
                    cell.font = bold_font
                    cell.fill = orange_fill
                    cell.border = border_style

    # Apply formatting to all sheets
    for sheet_name in sheets_to_format:
        sheet = wb[sheet_name]
        apply_comma_format(sheet)
        autofit_columns(sheet)
        format_total_rows(sheet)
        format_grand_total_rows(sheet)

    # Save the workbook to the in-memory file buffer
    output_file.seek(0)  # Reset buffer
    output_file = io.BytesIO()  # Create a fresh buffer to save the modified workbook
    wb.save(output_file)

    # Reset the buffer before sending for download
    output_file.seek(0)

    # Store the file in the Flask app config to be accessed later
    current_app.config['EXCEL_FILE'] = {
        'name': 'CY & LY Comparison.xlsx',
        'data': output_file.getvalue()}

    # Redirect to the processing page
    return redirect(url_for('processing'))

@cyly.route('/processing')
def processing():
    return render_template('processing.html')

@cyly.route('/download')
def download_report():
    # Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])

@cyly.route('/cyly-master', methods=['GET'])
def cyly_master():
    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\ITEM DESCRIPTION WITH CATEGORY (map).xlsx'
    return send_file(default_allocation_path, as_attachment=True)

@cyly.route('/cyly-lastyr', methods=['GET'])
def cyly_lastyr():
    # Path to the default lastyear sales file
    default_lastyr_path = r'D:\2MJIPL_COIMBATORE\Sales Register April23 to March24.xlsx' 
    return send_file(default_lastyr_path, as_attachment=True)

if __name__ == '__main__':
    cyly.run(debug=True)