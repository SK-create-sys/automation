from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
import re
import numpy as np
import openpyxl as px
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import xlsxwriter
from tempfile import NamedTemporaryFile
from werkzeug.utils import secure_filename
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Border, Side,Alignment

expnsmvmt = Blueprint('expnsmvmt', __name__)

@expnsmvmt.route('/')
def index():
    return render_template('test.html')

@expnsmvmt.route('/expnsmvmt-upload', methods=['POST'])
def upload_files():
        # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\master of Expenses.xlsx'
    # Check if an allocation file was provided by the user
    allocation_file = request.files.get('allocation')
    if allocation_file:
        # If the user provided an allocation file, delete the existing default file
        if os.path.exists(default_allocation_path):
            os.remove(default_allocation_path)  # Delete the existing default allocation file
        allocation_file.save(default_allocation_path)  # Save the new allocation file
        master_df = pd.read_excel(allocation_file, sheet_name = 'Gl master')
        factory_sales_df = pd.read_excel(allocation_file, sheet_name = 'factory & Sales')
    else:
        # If no allocation file is provided, use the default allocation file
        master_df = pd.read_excel(default_allocation_path, sheet_name = 'Gl master')
        factory_sales_df = pd.read_excel(default_allocation_path, sheet_name = 'factory & Sales')

# Read the uploaded Excel files
    trial_path = request.files['trial']
    trail_df = pd.read_excel(trial_path)

    trial_clean = trail_df[['Branch','Account No', 'Voucher Date','Departments','Profit & Loss']]

    for_cbe = ['CBE Print', 'CBEUS', 'Coimbatore Common Branch', 'Coimbatore Domestic Branch', 'Coimbatore Export Branch']
    for_kol_eou = ['Kolkata Common Branch', 'Kolkata EOU Branch', 'Shared Export Branch']
    for_kol = ['Guwahati Branch', 'Kolkata DOM Branch']

    def allocate_branch(row):
        branch = row['Branch']
        department = row['Departments']
        
        # Check for 'CBE' or 'KOL EOU'
        if branch in for_cbe:
            return 'CBE'
        elif branch in for_kol_eou:
            return 'KOL EOU'
        # Check for 'RED' condition
        elif branch in for_kol and (department is not None and not pd.isna(department)) and not (
            department.lower().startswith(('red', 'yellow')) or 
            department in ['Export Africa Sales', 'Export AUS/EUR Sales', 'Other Export Sales']):
            return 'KOL DOM'
        # Check for 'RED' condition based on Departments
        elif department is not None and not pd.isna(department) and department.lower().startswith('red'):
            return 'RED'
        elif department is not None and not pd.isna(department) and department.lower().startswith('yellow'):
            return 'YELLOW'
        else:
            return np.nan  # Return NaN for non-matching branches

    # Create the Allocated Branch column
    trial_clean['Allocated Branch'] = trial_clean.apply(allocate_branch, axis=1)

    added_master = pd.merge(trial_clean,master_df, left_on = 'Account No',right_on = 'No.', how = 'left')
    added_master_clean = added_master[(added_master['Account Type'].isin(['Posting'])) & (added_master['Account Category'].isin(['Expense']))]

    add_sales_factory_master = pd.merge(added_master_clean,factory_sales_df, on = ['Branch', 'Departments'], how = 'left')

    error_df =add_sales_factory_master[['Branch','Departments','Allocation']]
    error_file = error_df[(error_df['Allocation'].isna())]
    error_file = error_file.drop_duplicates()

    # add_sales_factory_master.loc[: ,'Month'] = add_sales_factory_master['Voucher Date'].dt.strftime('%B-%y')



    # Function to create pivot table and calculate YTD for a specific branch
    def create_pivot_and_ytd(df):
        df.loc[:, 'Month'] = df['Voucher Date'].dt.strftime('%B-%y')

        # Get unique available months
        available_months = df['Month'].unique().tolist()
        available_months_sorted = sorted(available_months, key=lambda x: pd.to_datetime(x, format='%B-%y'))

        # Create the pivot table using the available months
        pivot_table = df.pivot_table(
            index=['Allocation', 'Account Subcategory'],
            columns='Month',
            values='Profit & Loss',
            aggfunc='sum',
            fill_value=0
        ).reset_index()

        # Reindex the pivot table columns to include only the dynamically sorted available months
        pivot_table = pivot_table.reindex(columns=['Allocation', 'Account Subcategory'] + available_months_sorted)

        # Initialize the YTD column
        pivot_table['YTD 24-25'] = 0

        # Calculate YTD row-wise
        for month in available_months_sorted:
            pivot_table['YTD 24-25'] += pivot_table[month]

        # Calculate totals
        totals = pivot_table.groupby(['Allocation']).sum(numeric_only=True).reset_index()
        totals['Account Subcategory'] = '' 
        totals['Allocation'] = totals['Allocation'] + ' - Total'

        # Combine the pivot table and totals
        pivot_with_totals = pd.concat([pivot_table, totals], ignore_index=True)

        # Sort the DataFrame to ensure "Total" rows come last within each Allocation
        pivot_with_totals['sort_order'] = pivot_with_totals.apply(lambda row: (row['Allocation'].replace(' - Total', ''), 1) if 'Total' in row['Allocation'] else (row['Allocation'], 0), axis=1)
        pivot_with_totals = pivot_with_totals.sort_values(by='sort_order').drop(columns='sort_order')

        # Reset the index to keep it clean
        pivot_with_totals = pivot_with_totals.reset_index(drop=True)

        return pivot_with_totals

    # Function to add grouping in Excel
    def add_grouping(ws: Worksheet):
        start_row = None
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            allocation_value = row[0]  # Check the 'Allocation' column
            
            # Detect the '- Total' rows
            if allocation_value and '- Total' in allocation_value:
                if start_row is not None:  # If there's an ongoing group, close it
                    ws.row_dimensions.group(start_row, row_idx - 1, hidden=False)
                start_row = None  # Reset the start row after closing the group
            else:
                if start_row is None:  # Mark the start of the new group
                    start_row = row_idx
    # Function to apply styling to the worksheet
    def style_worksheet(ws: Worksheet):
        # Define styles
        bold_font = Font(bold=True, name='Calibri')
        body_font = Font(name='Calibri')
        purple_fill = PatternFill(start_color='BDA0E2', end_color='BDA0E2', fill_type='solid')  # Purple, Accent4, Lighter60%
        white_fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')  # White, Background1, Darker5%
        light_orange_fill = PatternFill(start_color='FFB06B', end_color='FFB06B', fill_type='solid')  # Orange, Accent6, Lighter40%
        center_alignment = Alignment(horizontal='center')

        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                            top=Side(style='thin'), bottom=Side(style='thin'))

        # Apply styles to the header row
        for cell in ws[1]:
            cell.font = bold_font
            cell.fill = purple_fill
            cell.alignment = center_alignment  # Center the header content

        # Get the indices for 'Allocation' and 'Account Subcategory' columns
        allocation_col_idx = None
        account_subcategory_col_idx = None
        for col_idx, cell in enumerate(ws[1], start=1):
            if cell.value == 'Allocation':
                allocation_col_idx = col_idx
            elif cell.value == 'Account Subcategory':
                account_subcategory_col_idx = col_idx

        # Apply styles to the rows
        for row in ws.iter_rows(min_row=2):
            allocation_cell = row[allocation_col_idx - 1]  # 'Allocation' column
            account_subcategory_cell = row[account_subcategory_col_idx - 1]  # 'Account Subcategory' column
            
            # Set background fill color for 'Allocation' and 'Account Subcategory' columns row values
            allocation_cell.fill = white_fill
            account_subcategory_cell.fill = white_fill

            # Set comma style and format with a thousands separator
            for cell in row[2:]:  # Assuming the first two columns are 'Allocation' and 'Account Subcategory'
                if isinstance(cell.value, (int, float)):  # Check if it's a number
                    cell.number_format = '#,##0.00'  # Set comma style
                cell.border = thin_border  # Apply borders

            # Check for Total rows and apply specific formatting
            if ' - Total' in allocation_cell.value:
                for cell in row:
                    cell.fill = light_orange_fill# Apply fill color for Total rows

        # Set the color for YTD column values
        ytd_col_idx = len(row) - 1  # Assuming 'YTD 24-25' is the last column
        for row in ws.iter_rows(min_row=2):
            ytd_cell = row[ytd_col_idx]
            ytd_cell.fill = light_orange_fill  # Apply fill color for YTD column

        # Apply borders to all cells
        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin_border

        # Autofit column widths
        for column in ws.columns:
            max_length = 0
            column = [cell for cell in column]
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)  # Add some padding
            ws.column_dimensions[column[0].column_letter].width = adjusted_width

    # Create an in-memory output file
    output_file = io.BytesIO()

    # Create an Excel workbook
    wb = Workbook()

    # List of allocated branches to process
    allocated_branches = ['KOL DOM', 'KOL EOU', 'CBE', 'RED', 'YELLOW']

    # Loop through each allocated branch and create the pivot tables
    for branch in allocated_branches:
        # Filter the DataFrame for the current branch
        branch_df = add_sales_factory_master[add_sales_factory_master['Allocated Branch'] == branch]
        
        # Create the pivot table for the current branch
        pivot_table = create_pivot_and_ytd(branch_df)

        # Create a new worksheet for this branch
        ws = wb.create_sheet(title=branch)

        # Write the DataFrame into the Excel worksheet
        for r_idx, row in enumerate(dataframe_to_rows(pivot_table, index=False, header=True), start=1):
            ws.append(row)

        # Add the grouping for rows based on the '- Total' rows
        add_grouping(ws)

        # Style the worksheet according to the specifications
        style_worksheet(ws)

    # Remove the default sheet created with the workbook
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Function to add DataFrame to a new worksheet
    def add_dataframe_to_sheet(df: pd.DataFrame, sheet_name: str, wb: Workbook):
        ws = wb.create_sheet(title=sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), start=1):
            ws.append(row)

    # Add the 'Error_Rows' sheet with the error_file DataFrame as the last sheet
    add_dataframe_to_sheet(added_master , "rawFile" , wb)
    add_dataframe_to_sheet(error_file, "Error_Rows", wb)

        # ** Save the workbook to the in-memory output_file **
    wb.save(output_file)

    # ** Reset the buffer position to the beginning of the file **
    output_file.seek(0)

    # Prepare the in-memory file for download
    output_file.seek(0)  # Reset the buffer before sending for download
    current_app.config['EXCEL_FILE'] = {
        'name': 'Expenses Wise Movement Report.xlsx',
        'data': output_file.getvalue()}
    
    # Redirect to the processing page
    return redirect(url_for('processing'))

@expnsmvmt.route('/processing')
def processing():
    return render_template('processing.html')

@expnsmvmt.route('/download')
def download_report():
# Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'],mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

@expnsmvmt.route('/expnsmvmt-master', methods=['GET'])
def insurance_master():
    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\master of Expenses.xlsx'
    return send_file(default_allocation_path, as_attachment=True)

if __name__ == '__main__':
    expnsmvmt.run(debug=True)

