from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
import re
import numpy as np
import openpyxl as px
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import xlsxwriter
import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl import load_workbook
from werkzeug.utils import secure_filename

ecomarket = Blueprint('ecomarket', __name__)

UPLOAD_FOLDER = r'D:\2MJIPL_COIMBATORE\Overall Sales Report'


# Route to display available months and file upload form
@ecomarket.route("/", methods=["GET"])
def index():
    return render_template('test.html')

# Route for handling file uploads and processing
@ecomarket.route('/ecomarket-upload', methods=['POST'])
def upload_file():

    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\Overall Sales Report\Master.xlsx'
    # Check if an allocation file was provided by the user
    allocation_file = request.files.get('allocation')
    if allocation_file:
        # If the user provided an allocation file, delete the existing default file
        if os.path.exists(default_allocation_path):
            os.remove(default_allocation_path)  # Delete the existing default allocation file
        allocation_file.save(default_allocation_path)  # Save the new allocation file
        master = pd.read_excel(allocation_file, sheet_name = 'Amazon')
        allocation = pd.read_excel(allocation_file)
    else:
        # If no allocation file is provided, use the default allocation file
        master = pd.read_excel(default_allocation_path, sheet_name = 'Amazon')
        allocation = pd.read_excel(default_allocation_path)

    amazon_market_place_file = request.files['amazon_market_place']  
    amazon_retail_file = request.files['ARIPL_Sales']  
    bigbasket_market_place_file = request.files['BigBasket_Sales']
    flipkart_shopsy_market_place_file = request.files['Flipkart&Shopsy_Sales']
    flipkart_grocery_file = request.files['flipkart_grocery']
    JioMart_file =  request.files['JioMart_Sales']
    Swiggy_Instamart_file = request.files['Swiggy_Sales']

        # Save uploaded files to the server
    if amazon_market_place_file:
        amazon_market_place_filename = secure_filename(amazon_market_place_file.filename)
        amazon_market_place_path = os.path.join(UPLOAD_FOLDER, "Amazon Marketplace Sales", amazon_market_place_filename)
        os.makedirs(os.path.dirname(amazon_market_place_path), exist_ok=True)
        amazon_market_place_file.save(amazon_market_place_path)

    if amazon_retail_file:
        amazon_retail_filename = secure_filename(amazon_retail_file.filename)
        amazon_retail_path = os.path.join(UPLOAD_FOLDER, "ARIPL Sales", amazon_retail_filename)
        os.makedirs(os.path.dirname(amazon_retail_path), exist_ok=True)
        amazon_retail_file.save(amazon_retail_path)

    if bigbasket_market_place_file:
        bigbasket_market_place_filename = secure_filename(bigbasket_market_place_file.filename)
        bigbasket_market_place_path = os.path.join(UPLOAD_FOLDER, "BigBasket Sales", bigbasket_market_place_filename)
        os.makedirs(os.path.dirname(bigbasket_market_place_path), exist_ok=True)
        bigbasket_market_place_file.save(bigbasket_market_place_path)

    if flipkart_shopsy_market_place_file:
        flipkart_shopsy_market_place_filename = secure_filename(flipkart_shopsy_market_place_file.filename)
        flipkart_shopsy_market_place_path = os.path.join(UPLOAD_FOLDER, "Flipkart & Shopsy Marketplace Sales", flipkart_shopsy_market_place_filename)
        os.makedirs(os.path.dirname(flipkart_shopsy_market_place_path), exist_ok=True)
        flipkart_shopsy_market_place_file.save(flipkart_shopsy_market_place_path)   


    if flipkart_grocery_file:
        flipkart_grocery_filename = secure_filename(flipkart_grocery_file.filename)
        flipkart_grocery_path = os.path.join(UPLOAD_FOLDER, "Flipkart Grocery Sales", flipkart_grocery_filename)
        os.makedirs(os.path.dirname(flipkart_grocery_path), exist_ok=True)
        flipkart_grocery_file.save(flipkart_grocery_path)          

    if JioMart_file:
        JioMart_filename = secure_filename(JioMart_file.filename)
        JioMart_path = os.path.join(UPLOAD_FOLDER, "JioMart", JioMart_filename)
        os.makedirs(os.path.dirname(JioMart_path), exist_ok=True)
        JioMart_file.save(JioMart_path)       

    if Swiggy_Instamart_file:
        Swiggy_Instamart_filename = secure_filename(Swiggy_Instamart_file.filename)
        Swiggy_Instamart_path = os.path.join(UPLOAD_FOLDER, "Swiggy Instamart", Swiggy_Instamart_filename)
        os.makedirs(os.path.dirname(Swiggy_Instamart_path), exist_ok=True)
        Swiggy_Instamart_file.save(Swiggy_Instamart_path)          


    #AMAZON MARKET PLACE
    amazon_market_place  = os.path.join(UPLOAD_FOLDER,"Amazon Marketplace Sales")
    market_place_df = []
    for filename in os.listdir(amazon_market_place):
        if filename.endswith('.csv'):
            month = os.path.splitext(filename)[0]
            file_path = os.path.join(amazon_market_place, filename)
            amp_df = pd.read_csv(file_path)
            amp_df['Month'] = month
            amp_df['Ordered Product Sales'] = amp_df['Ordered Product Sales'].replace('[₹,]', '', regex=True).astype(float).astype(int)
            market_place_df.append(amp_df)

    AMP_merged = pd.concat(market_place_df, ignore_index=True)
    inner_AMP = pd.merge(AMP_merged, master, left_on='(Parent) ASIN', right_on='Amazon ASIN', how='inner')

    inner_AMP['Units'] = inner_AMP['No of Boxes']* inner_AMP['Units Ordered']
    inner_AMP = inner_AMP.rename(columns={'Ordered Product Sales': 'Amounts'})
    #final
    amazon_market_final_data = inner_AMP[['Item ID','Month','Units','Amounts']]

    blank_AMP = pd.merge(AMP_merged, master, left_on='(Parent) ASIN', right_on='Amazon ASIN', how='left')
    filtered_blank_AMP = blank_AMP[blank_AMP['Amazon ASIN'].isna()]
    #blank ASIN
    filtered_blank_AMP = filtered_blank_AMP[['(Parent) ASIN']].drop_duplicates()



    #####################################################################################################################################################

    #AMAZON RETAIL (ARIPL)
    amazon_retail  = os.path.join(UPLOAD_FOLDER,"ARIPL Sales")
    amazon_retail_df = []
    for ar_name in os.listdir(amazon_retail):
        if ar_name.endswith('.xlsx'):
            month = os.path.splitext(ar_name)[0]
            amazon_retail_path = os.path.join(amazon_retail,ar_name)
            ar_df = pd.read_excel(amazon_retail_path , skiprows = 1)
            ar_df['Month'] = month
            amazon_retail_df.append(ar_df)

    ar_merged = pd.concat(amazon_retail_df, ignore_index=True)
    inner_ar = pd.merge(ar_merged, master, left_on='ASIN', right_on='Amazon ASIN', how='inner')

    inner_ar['Units'] = inner_ar['No of Boxes'] * inner_ar['Shipped Units']
    inner_ar = inner_ar.rename(columns={'Shipped COGS': 'Amounts'})
    #final
    amazon_ar_final_data =inner_ar[['Item ID','Month','Units','Amounts']]

    blank_ar = pd.merge(ar_merged, master, left_on='ASIN', right_on='Amazon ASIN', how='left')
    filtered_blank_ar = blank_ar[blank_ar['Amazon ASIN'].isna()]
    #blank ASIN
    filtered_blank_ar = filtered_blank_ar[['ASIN']].drop_duplicates()

    # filtered_blank_ar.to_excel("D:\\Overall Sales Report\\Test2.xlsx")
    ######################################################################################################################################################


    #BIGBASKET SALES
    bigbasket_market_place  = os.path.join(UPLOAD_FOLDER,"BigBasket Sales")
    bigbasket_market_place_df = []
    for filename in os.listdir(bigbasket_market_place):
        if filename.endswith('.csv'):
            month = os.path.splitext(filename)[0]
            file_path = os.path.join(bigbasket_market_place, filename)
            bb_df = pd.read_csv(file_path)
            bb_df['Month'] = month
            bigbasket_market_place_df.append(bb_df)

    bb_merged = pd.concat(bigbasket_market_place_df, ignore_index=True)

    bb_merged['source_sku_id'] = bb_merged['source_sku_id'].astype(str)
    master_bb = master.dropna(subset=['BigBasket'])
    master_bb = master_bb[pd.to_numeric(master_bb['BigBasket'], errors='coerce').notnull()]


    master_bb['BigBasket'] = master_bb['BigBasket'].astype(float).astype(int).astype(str)


    inner_bb = pd.merge(bb_merged, master_bb, left_on='source_sku_id', right_on='BigBasket', how='inner')

    inner_bb['Units'] = inner_bb['total_quantity']* inner_bb['No of Boxes']
    inner_bb = inner_bb.rename(columns={'total_sales': 'Amounts'})
    #final
    bigbasket_market_final_data = inner_bb[['Item ID','Month','Units','Amounts']]

    blank_bb = pd.merge(bb_merged, master_bb, left_on='source_sku_id', right_on='BigBasket', how='left')
    filtered_blank_bb = blank_bb[blank_bb['BigBasket'].isna()]
    #blank ASIN
    filtered_blank_bb = filtered_blank_bb[['source_sku_id']].drop_duplicates()


    ######################################################################################################################################################

    #FLIPKART & SHOPSY SALES
    flipkart_shopsy_market_place  = os.path.join(UPLOAD_FOLDER,"Flipkart & Shopsy Marketplace Sales")
    flipkart_shopsy_market_place_df = []
    for filename in os.listdir(flipkart_shopsy_market_place):
        if filename.endswith('.xlsx'):
            month = os.path.splitext(filename)[0]
            file_path = os.path.join(flipkart_shopsy_market_place, filename)
            FS_df = pd.read_excel(file_path, sheet_name = 'Sales Report')
            FS_df['Month'] = month
            FS_df['FSN'] = FS_df['FSN'].str.replace('"', '')
            flipkart_shopsy_market_place_df.append(FS_df)

    FS_merged = pd.concat(flipkart_shopsy_market_place_df, ignore_index=True)
    FS_raw_data = FS_merged[FS_merged['Event Type']== 'Sale']


    #For Flipkart
    flipkart_data = FS_raw_data[~FS_raw_data['FSN'].str.startswith('Y')]

    inner_flkt = pd.merge(flipkart_data, master, left_on='FSN', right_on='Flipkart FSN', how='inner')

    inner_flkt['Units'] = inner_flkt['Item Quantity']* inner_flkt['No of Boxes']
    inner_flkt = inner_flkt.rename(columns={'Price before discount': 'Amounts'})
    #final
    FLIPKART_final_data = inner_flkt[['Item ID','Month','Units','Amounts']]

    blank_flkt = pd.merge(flipkart_data, master, left_on='FSN', right_on='Flipkart FSN', how='left')
    filtered_blank_flkt = blank_flkt[blank_flkt['Flipkart FSN'].isna()]
    #blank ASIN
    filtered_blank_flkt = filtered_blank_flkt[['FSN']].drop_duplicates()


    #For SHOPSY
    Shopsy_data = FS_raw_data[FS_raw_data['FSN'].str.startswith('Y')]
    Shopsy_data=Shopsy_data.rename(columns={'FSN':'SHOPSY FSN'})

    inner_shopsy = pd.merge(Shopsy_data, master, left_on='SHOPSY FSN', right_on='Shopsy FSN', how='inner')

    inner_shopsy['Units'] = inner_shopsy['Item Quantity']* inner_shopsy['No of Boxes']
    inner_shopsy = inner_shopsy.rename(columns={'Price before discount': 'Amounts'})
    #final
    SHOPSY_final_data = inner_shopsy[['Item ID','Month','Units','Amounts']]

    blank_shopsy = pd.merge(Shopsy_data, master, left_on='SHOPSY FSN', right_on='Shopsy FSN', how='left')
    filtered_blank_shopsy = blank_shopsy[blank_shopsy['Shopsy FSN'].isna()]
    #blank ASIN
    filtered_blank_shopsy = filtered_blank_shopsy[['SHOPSY FSN']].drop_duplicates()


    #######################################################################################################################################################

    #FLIPKART GROCERY
    grocery_market_place  = os.path.join(UPLOAD_FOLDER,"Flipkart Grocery Sales")
    grocery_market_place_df = []
    for filename in os.listdir(grocery_market_place):
        if filename.endswith('.xlsx'):
            month = os.path.splitext(filename)[0]
            file_path = os.path.join(grocery_market_place, filename)
            fg_df = pd.read_excel(file_path, skiprows = 2)
            fg_df['Month'] = month
            grocery_market_place_df.append(fg_df)

    fg_merged = pd.concat(grocery_market_place_df, ignore_index=True)

    inner_fg = pd.merge(fg_merged, master, left_on='FSN', right_on='Flipkart FSN', how='inner')

    inner_fg['Units'] = inner_fg['Units']* inner_fg['No of Boxes']
    inner_fg = inner_fg.rename(columns={'GMV': 'Amounts'})
    inner_fg['Units'] = pd.to_numeric(inner_fg['Units'], errors='coerce')
    inner_fg['Amounts'] = pd.to_numeric(inner_fg['Amounts'], errors='coerce')
    #final
    grocery_market_final_data = inner_fg[['Item ID','Month','Units','Amounts']]

    blank_fg = pd.merge(fg_merged, master, left_on='FSN', right_on='Flipkart FSN', how='left')
    filtered_blank_fg = blank_fg[blank_fg['Flipkart FSN'].isna()]
    #blank ASIN
    filtered_blank_fg = filtered_blank_fg[['FSN']].drop_duplicates()


    ######################################################################################################################################################

    #JIO-MART 
    jiomart_path  = os.path.join(UPLOAD_FOLDER,"JioMart")
    jiomart_df = []
    for jio_name in os.listdir(jiomart_path):
        if jio_name.endswith('.csv'):
            month = os.path.splitext(jio_name)[0]
            jiomart_retail_path = os.path.join(jiomart_path,jio_name)
            jio_df = pd.read_csv(jiomart_retail_path)
            jio_df['Month'] = month
            jiomart_df.append(jio_df)
    df_list = [df.dropna(how='all') for df in jiomart_df] 
    jio_merged = pd.concat(df_list, ignore_index=True)

    master['JioMart_2'] = master['JioMart_2'].astype(str)
    jio_merged['SKU'] = jio_merged['SKU'].astype(str)

    inner_jio = pd.merge(jio_merged, master, left_on='SKU', right_on='JioMart_2', how='inner')

    inner_jio['Units'] = inner_jio['No of Boxes'] * inner_jio['Item Quantity']
    inner_jio = inner_jio.rename(columns={'Final Invoice Amount': 'Amounts'})
    #final
    jiomart_final_data =inner_jio[['Item ID','Month','Units','Amounts']]

    blank_jio = pd.merge(jio_merged, master, left_on='SKU', right_on='JioMart_2', how='left')
    filtered_blank_jio = blank_jio[blank_jio['JioMart_2'].isna()]
    #blank ASIN
    filtered_blank_jio = filtered_blank_jio[['SKU']].drop_duplicates()


    ######################################################################################################################################################

    #SWIGGY INSTAMART
    swiggy_path  = os.path.join(UPLOAD_FOLDER,"Swiggy Instamart")
    swiggy_df = []
    for swiggy_name in os.listdir(swiggy_path):
        if swiggy_name.endswith('.xlsx'):
            month = os.path.splitext(swiggy_name)[0]
            swiggy_retail_path = os.path.join(swiggy_path,swiggy_name)
            swgy_df = pd.read_excel(swiggy_retail_path)
            swgy_df['Month'] = month
            swiggy_df.append(swgy_df)

    swiggy_merged = pd.concat(swiggy_df, ignore_index=True)

    master = master[pd.to_numeric(master['Swiggy Instamart'], errors='coerce').notnull()]
    master['Swiggy Instamart'] = master['Swiggy Instamart'].astype(float).astype(int).astype(str)
    swiggy_merged['FINAL_ITEM_CODE'] = swiggy_merged['FINAL_ITEM_CODE'].astype(float).astype(int).astype(str)

    inner_swgy = pd.merge(swiggy_merged, master, left_on='FINAL_ITEM_CODE', right_on='Swiggy Instamart', how='inner')

    inner_swgy['Units'] = inner_swgy['No of Boxes'] * inner_swgy['FINAL_QTY']
    inner_swgy = inner_swgy.rename(columns={'FINAL_GMV': 'Amounts'})
    #final
    swiggy_final_data =inner_swgy[['Item ID','Month','Units','Amounts']]

    blank_swiggy = pd.merge(swiggy_merged, master, left_on='FINAL_ITEM_CODE', right_on='Swiggy Instamart', how='left')
    filtered_blank_swiggy = blank_swiggy[blank_swiggy['Swiggy Instamart'].isna()]
    #blank ASIN
    filtered_blank_swiggy = filtered_blank_swiggy[['FINAL_ITEM_CODE']].drop_duplicates()


    #######################################################################################################################################################

    # making master output
    concat_final_files = pd.concat([amazon_ar_final_data,amazon_market_final_data,bigbasket_market_final_data,FLIPKART_final_data,SHOPSY_final_data,grocery_market_final_data,jiomart_final_data,swiggy_final_data], ignore_index=True)
    concat_blank_files = pd.concat([filtered_blank_ar,filtered_blank_AMP,filtered_blank_bb,filtered_blank_flkt,filtered_blank_shopsy,filtered_blank_fg,filtered_blank_jio,filtered_blank_swiggy], ignore_index=True)
    allocation = pd.read_excel(r'D:\Overall Sales Report\Master.xlsx')
    output_raw_with_allocation= pd.merge(concat_final_files,allocation, on = 'Item ID', how= 'inner')


    ######################################################################################################################################################

    month_order = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 
            'February', 'March']

    variable_order = ['Units', 'Amounts']

    month_mapping = {month: idx for idx, month in enumerate(month_order)}
    variable_mapping = {variable: idx for idx, variable in enumerate(variable_order)}

    melted_df =output_raw_with_allocation.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0)

    sorted_columns = sorted(output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    output = output[sorted_columns]

    total_units= output.filter(like='Units').sum(axis=1)
    total_amounts = output.filter(like='Amounts').sum(axis=1)
    output['Total', 'Units'] = total_units
    output['Total', 'Amounts'] = total_amounts
    Gross_Total= output['Total', 'Amounts'].sum()

    output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_output = output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)


    #####################################################################################################################################################


    AmazonMarket= pd.merge(amazon_market_final_data,allocation, on = 'Item ID', how= 'inner')

    melted_df =AmazonMarket.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    amazon_market_final_data_output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0)

    sorted_columns = sorted(amazon_market_final_data_output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    amazon_market_final_data_output = amazon_market_final_data_output[sorted_columns]

    tota_units= amazon_market_final_data_output.filter(like='Units').sum(axis=1)
    total_amounts = amazon_market_final_data_output.filter(like='Amounts').sum(axis=1)
    amazon_market_final_data_output['Total', 'Units'] = tota_units
    amazon_market_final_data_output['Total', 'Amounts'] = total_amounts
    Gross_Total= amazon_market_final_data_output['Total', 'Amounts'].sum()

    amazon_market_final_data_output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_amazon_market_final_data_output= amazon_market_final_data_output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)
    # sorted_amazon_market_final_data_output-------------------------------------------

    AmazonRetail= pd.merge(amazon_ar_final_data,allocation, on = 'Item ID', how= 'inner')

    melted_df =AmazonRetail.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    amazon_ar_final_data_output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0)

    sorted_columns = sorted(amazon_ar_final_data_output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    amazon_ar_final_data_output =amazon_ar_final_data_output[sorted_columns]

    tota_units= amazon_ar_final_data_output.filter(like='Units').sum(axis=1)
    total_amounts = amazon_ar_final_data_output.filter(like='Amounts').sum(axis=1)
    amazon_ar_final_data_output['Total', 'Units'] = tota_units
    amazon_ar_final_data_output['Total', 'Amounts'] = total_amounts
    Gross_Total= amazon_ar_final_data_output['Total', 'Amounts'].sum()

    amazon_ar_final_data_output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_amazon_ar_final_data_output= amazon_ar_final_data_output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)
    # sorted_amazon_ar_final_data_output-------------------------------------------

    BigBasket= pd.merge(bigbasket_market_final_data,allocation, on = 'Item ID', how= 'inner')

    melted_df =BigBasket.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    bigbasket_market_final_data_output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0)

    sorted_columns = sorted(bigbasket_market_final_data_output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    bigbasket_market_final_data_output =bigbasket_market_final_data_output[sorted_columns]

    tota_units= bigbasket_market_final_data_output.filter(like='Units').sum(axis=1)
    total_amounts = bigbasket_market_final_data_output.filter(like='Amounts').sum(axis=1)
    bigbasket_market_final_data_output['Total', 'Units'] = tota_units
    bigbasket_market_final_data_output['Total', 'Amounts'] = total_amounts
    Gross_Total= bigbasket_market_final_data_output['Total', 'Amounts'].sum()

    bigbasket_market_final_data_output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_bigbasket_market_final_data_output= bigbasket_market_final_data_output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)
    # sorted_bigbasket_market_final_data_output----------------------------------------

    flipkart_shopsy_data = pd.concat([FLIPKART_final_data,SHOPSY_final_data])
    flipkarshopsy= pd.merge(flipkart_shopsy_data,allocation, on = 'Item ID', how= 'inner')

    melted_df =flipkarshopsy.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    flipkarshopsy_data_output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0)

    sorted_columns = sorted(flipkarshopsy_data_output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    flipkarshopsy_data_output =flipkarshopsy_data_output[sorted_columns]

    tota_units= flipkarshopsy_data_output.filter(like='Units').sum(axis=1)
    total_amounts = flipkarshopsy_data_output.filter(like='Amounts').sum(axis=1)
    flipkarshopsy_data_output['Total', 'Units'] = tota_units
    flipkarshopsy_data_output['Total', 'Amounts'] = total_amounts
    Gross_Total= flipkarshopsy_data_output['Total', 'Amounts'].sum()

    flipkarshopsy_data_output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_flipkarshopsy_data_output= flipkarshopsy_data_output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)
    # sorted_flipkarshopsy_data_output-------------------------------------------

    FlipkartGrocery= pd.merge(grocery_market_final_data,allocation, on = 'Item ID', how= 'inner')

    melted_df =FlipkartGrocery.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    grocery_market_final_data_output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0)

    sorted_columns = sorted(grocery_market_final_data_output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    grocery_market_final_data_output = grocery_market_final_data_output[sorted_columns]

    total_units= grocery_market_final_data_output.filter(like='Units').sum(axis=1)
    total_amounts = grocery_market_final_data_output.filter(like='Amounts').sum(axis=1)
    grocery_market_final_data_output['Total', 'Units'] = total_units
    grocery_market_final_data_output['Total', 'Amounts'] = total_amounts
    Gross_Total= grocery_market_final_data_output['Total', 'Amounts'].sum()

    grocery_market_final_data_output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_grocery_market_final_data_output= grocery_market_final_data_output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)
    #sorted_Grocery_final_data_output--------------------------

    JioMart = pd.merge(jiomart_final_data,allocation, on = 'Item ID', how= 'inner')

    melted_df =JioMart.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    jiomart_final_data_output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0) 

    sorted_columns = sorted(jiomart_final_data_output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    jiomart_final_data_output = jiomart_final_data_output[sorted_columns]

    total_units= jiomart_final_data_output.filter(like='Units').sum(axis=1)
    total_amounts = jiomart_final_data_output.filter(like='Amounts').sum(axis=1)
    jiomart_final_data_output['Total', 'Units'] = total_units
    jiomart_final_data_output['Total', 'Amounts'] = total_amounts
    Gross_Total= jiomart_final_data_output['Total', 'Amounts'].sum()

    jiomart_final_data_output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_jiomart_final_data_output= jiomart_final_data_output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)
    #sorted_jiomart_final_data_output


    SwiggyInsta = pd.merge(swiggy_final_data,allocation, on = 'Item ID', how= 'inner')

    melted_df =SwiggyInsta.melt(id_vars=['Item ID' ,'Brand Name', 'SS Name', 'Category', 'Month'], var_name='Variable', value_name='Metrics')

    swiggy_final_data_output = pd.pivot_table(melted_df,
                            index=['Item ID' ,'Brand Name', 'SS Name', 'Category'],
                            columns=['Month','Variable'],
                            values='Metrics',
                            aggfunc='sum',
                            fill_value=0) 

    sorted_columns = sorted(swiggy_final_data_output.columns, key=lambda x: (month_order.index(x[0]), variable_order.index(x[1])))
    swiggy_final_data_output = swiggy_final_data_output[sorted_columns]

    total_units= swiggy_final_data_output.filter(like='Units').sum(axis=1)
    total_amounts = swiggy_final_data_output.filter(like='Amounts').sum(axis=1)
    swiggy_final_data_output['Total', 'Units'] = total_units
    swiggy_final_data_output['Total', 'Amounts'] = total_amounts
    Gross_Total= swiggy_final_data_output['Total', 'Amounts'].sum()

    swiggy_final_data_output['Base on Total Amount', 'Contribution (%)'] = round((total_amounts /Gross_Total) * 100,2)
    sorted_swiggy_final_data_output= swiggy_final_data_output.sort_values(by=('Base on Total Amount', 'Contribution (%)'), ascending=False)
    #sorted_jiomart_final_data_output


    output_file = io.BytesIO()

    with pd.ExcelWriter(output_file, engine='xlsxwriter') as writer:
        sorted_output.to_excel(writer, sheet_name='Combined Market Place', index=True)
        sorted_amazon_market_final_data_output.to_excel(writer, sheet_name='Amazon Marketplace Sales', index=True)
        sorted_amazon_ar_final_data_output.to_excel(writer, sheet_name='ARIPL Sales', index=True)
        sorted_bigbasket_market_final_data_output.to_excel(writer, sheet_name='BigBasket Sales', index=True)
        sorted_flipkarshopsy_data_output.to_excel(writer, sheet_name='Flipkart & Shopsy Sales', index=True)
        sorted_grocery_market_final_data_output.to_excel(writer, sheet_name='Flipkart Grocery', index=True)
        sorted_jiomart_final_data_output.to_excel(writer, sheet_name='JioMart', index=True)
        sorted_swiggy_final_data_output.to_excel(writer, sheet_name='Swiggy Instamart', index=True)
        concat_blank_files.to_excel(writer, sheet_name='Not in Master ASIN', index=False)
        

    # Save the modified workbook to the in-memory buffer
    output_file.seek(0)  # Reset the buffer before saving
    # new_output_file = io.BytesIO()  # Create a fresh buffer for the final output
    # new_output_file.seek(0) 

        # Store the file in the Flask app config
    current_app.config['EXCEL_FILE'] = {
        'name': 'Total Online Market Sales.xlsx',
        'data': output_file.getvalue()
    }   

    # Redirect to the processing page
    return redirect(url_for('processing'))


@ecomarket.route('/download')
def download_report():
     # Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])


@ecomarket.route('/ecomarket-master', methods=['GET'])
def ecomarket_master():
    # Path to the default lastyear sales file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\Overall Sales Report\Master.xlsx'
    return send_file(default_allocation_path, as_attachment=True)



from io import BytesIO
# Regular expression to extract the month name from the filename
month_pattern = re.compile(r'(April|May|June|July|August|September|October|November|December|January|February|March)', re.IGNORECASE)

# Financial year month order (April to March)
financial_month_order = ['April', 'May', 'June', 'July', 'August', 'September', 'October', 'November', 'December', 'January', 'February', 'March']


@ecomarket.route('/process-filenames', methods=['GET'])
def process_files():
    # Retrieve subfolders
    subfolders = [f.path for f in os.scandir(UPLOAD_FOLDER) if f.is_dir()]

    # Dictionary to hold subfolder names and their corresponding files
    subfolder_files = {}

    for subfolder in subfolders:
        # Get the subfolder name (without the full path)
        subfolder_name = os.path.basename(subfolder)
        try:
            # List all files in the subfolder
            files = os.listdir(subfolder)
            
            # Filter and match files by month names
            matched_files = []
            for file in files:
                match = month_pattern.search(file)
                if match:
                    matched_files.append(match.group(0).capitalize())  # Capitalize month name
            
            # Sort the matched files by financial year month order
            matched_files_sorted = sorted(matched_files, key=lambda x: financial_month_order.index(x))
            
            # Store the sorted files in the dictionary
            subfolder_files[subfolder_name] = matched_files_sorted

        except Exception as e:
            return f"An error occurred while processing files in '{subfolder_name}': {str(e)}", 500

    # If no subfolders found, return an error
    if not subfolder_files:
        return "No subfolders with matching files found.", 404

    # Create a DataFrame with subfolders as columns and months as rows
    max_rows = max(len(files) for files in subfolder_files.values())  # Get max number of rows (months)
    df = pd.DataFrame({subfolder: pd.Series(files) for subfolder, files in subfolder_files.items()})

    # Fill missing rows with empty values
    df = df.fillna('')

    # Write DataFrame to an Excel file in memory
    output = BytesIO()
    try:
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Available Months')
        output.seek(0)

        # Return the Excel file as an attachment for download
        return send_file(output, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                         as_attachment=True, download_name='Available_Months.xlsx')

    except Exception as e:
        return f"An error occurred while generating the Excel file: {str(e)}", 500




if __name__ == '__main__':
    ecomarket.run(debug=True)
