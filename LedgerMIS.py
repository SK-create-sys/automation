from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import pandas as pd
import re
import numpy as np
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side
import xlsxwriter
from tempfile import NamedTemporaryFile
from openpyxl import load_workbook
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Border, Side

ledger = Blueprint('ledger', __name__)

@ledger.route('/')
def index():
    return render_template('test.html')

@ledger.route('/ledger-upload', methods=['POST'])
def upload_files():
    ledger = request.files['ledger']
    sales = request.files['sales']

    # Read the uploaded Excel files
    file_path = pd.read_excel(ledger)
    sales_path = pd.read_excel(sales)

    remove_BA = file_path[(file_path['Ledger Groups'] != 'Balancing Accounts')]

    include_dimension = ['KOL DOM', 'KOL EOU','KOL COM', 'RED GEN', 'YEL GEN']
    filter_dimension = remove_BA[remove_BA['Dimension1'].isin(include_dimension)]
    include_dpt_for_dimension = [ 'Single Chamber', 'Constanta','IMA','Box Former','IT and System','Audit and Compliance','FFS','Perfecta',
                'Accounts and Finance','FFS less Than 50 Gms', 'FFS 51- 100 Gms', 'FFS 250 Gms- 1 KG','QA and QC', 
                'MD 20', 'BOPP', 'Administration and HR', 'Pyramid','YELLOW General','RED General']
    filter_dimension_with_dpt = filter_dimension[filter_dimension['Departments'].isin(include_dpt_for_dimension)]

    #

    include_next_dimension = ['RED UP','RED PUN', 'RED UK', 'RED BIH','RED RAJ', 'RED WB', 'PKT MAH', 'RED MP','RED JKD', 'RED HP']
    filter_next_dimension = remove_BA[remove_BA['Dimension1'].isin(include_next_dimension)]
    include_dpt_for_next_dimension = ['Single Chamber', 'Constanta','IMA','Box Former','FFS','Perfecta',
                'FFS less Than 50 Gms', 'FFS 51- 100 Gms', 'FFS 250 Gms- 1 KG', 
                'MD 20', 'BOPP', 'Pyramid','YELLOW General','RED General']
    filter_next_dimension_with_dpt = filter_next_dimension[filter_next_dimension['Departments'].isin(include_dpt_for_next_dimension)]

    #

    Dimension_department = filter_dimension_with_dpt.groupby(['Dimension1','Dimension2','Departments']).agg({'Profit & Loss': 'sum'}).reset_index()
    next_Dimension_department = filter_next_dimension_with_dpt.groupby(['Dimension1','Dimension2','Departments']).agg({'Profit & Loss': 'sum'}).reset_index()
    first_merged = pd.concat([Dimension_department,next_Dimension_department])

    #

    include_for_dimension = ['KOL DOM', 'KOL EOU','KOL COM','RED UP','RED PUN', 'RED UK', 'RED BIH','RED RAJ', 'RED WB', 'PKT MAH', 'RED MP',
                            'RED JKD', 'RED HP']
    filter_for_summary = first_merged[first_merged['Dimension1'].isin(include_for_dimension)]
    exclude_only_dpt = ['RED General','YELLOW General']
    filter_dimension_without_REDYLWGEN = filter_for_summary[~filter_for_summary['Departments'].isin(exclude_only_dpt)]
    summary_group_a = filter_dimension_without_REDYLWGEN.groupby(['Dimension1']).agg({'Profit & Loss': 'sum'}).reset_index()

    #

    include_only_dimension = ['RED GEN','YEL GEN']
    filter_dimension_only_REDYLWGEN = first_merged[first_merged['Dimension1'].isin(include_only_dimension)]
    REDYLWGEN_group = filter_dimension_only_REDYLWGEN.groupby(['Dimension1']).agg({'Profit & Loss': 'sum'}).reset_index()

    #                                             
                                                
    include_for_dimension = ['KOL DOM', 'KOL EOU','KOL COM','RED UP','RED PUN', 'RED UK', 'RED BIH','RED RAJ', 'RED WB', 'PKT MAH', 'RED MP',
                            'RED JKD', 'RED HP']
    filter_for_summary = first_merged[first_merged['Dimension1'].isin(include_for_dimension)]
    include_only_dpt = ['RED General','YELLOW General']
    filter_dimension_without_REDYLWGEN = filter_for_summary[filter_for_summary['Departments'].isin(include_only_dpt)]
    summary_group_b = filter_dimension_without_REDYLWGEN.groupby(['Departments']).agg({'Profit & Loss': 'sum'}).reset_index()

    summary_group_b.loc[summary_group_b['Departments'].str.contains(r'RED', case=False), 'Departments'] = 'RED GEN'
    summary_group_b.loc[summary_group_b['Departments'].str.contains(r'YELLOW', case=False), 'Departments'] = 'YEL GEN'
    summary_group_b=summary_group_b.rename(columns={'Departments':'Dimension1'})
    
    #

    concatenated_df = pd.concat([REDYLWGEN_group, summary_group_b])
    red_yellow = concatenated_df.groupby('Dimension1')['Profit & Loss'].sum().reset_index()
    second_merge_table = pd.concat([summary_group_a,red_yellow])
                                                
    #########################################################################################################################                                               
                                                
                                                
    sales_path['Invoice Date'] = pd.to_datetime(sales_path['Invoice Date'], errors='coerce')
    sales_path['Invoice Month'] = sales_path['Invoice Date'].dt.strftime('%B')
    # sales_path['Invoice Month'].unique() 
    # sales_path = sales_path[sales_path['Invoice Month'] == 'September']   #changing Slot

    #

    include_red_yellow_dept = ['RED Parwanoo', 'RED Zirakpur', 'RED Dehradun',
        'RED Jaipur', 'RED Lucknow', 'RED Ghaziabad','RED Gujarat','RED Patna','Red General Inst', 'RED Ranchi',
        'RED Indore', 'RED Jabalpur','YELLOW Pune', 'YELLOW Bangalore']
    red_yellow_df = sales_path[sales_path['Department'].isin(include_red_yellow_dept)]  
    red_yellow_group = red_yellow_df.groupby(['Department']).agg({'Sale Amount Actual': 'sum'}).reset_index()                                               

    #                                              
                                                
    red_df = red_yellow_group[red_yellow_group['Department'].str.contains(r'RED',case=False)]
    grand_total_red = red_df['Sale Amount Actual'].sum()
    total_red_profit_loss = second_merge_table[second_merge_table['Dimension1'] == 'RED GEN']['Profit & Loss'].values[0]
    grand_total_row = pd.DataFrame({'Department': ['Grand Total'], 'Sale Amount Actual': [grand_total_red], 'amnt': [total_red_profit_loss]})
    red_df = pd.concat([red_df, grand_total_row], ignore_index=True)
    red_df['amnt'] = (red_df['Sale Amount Actual'] * total_red_profit_loss) / grand_total_red
    # red_df = red_df.append({'Department': 'Grand Total', 'Sale Amount Actual':grand_total_red}, ignore_index=True)

    #                                               
                                                
    yellow_df = red_yellow_group[red_yellow_group['Department'].str.contains(r'YELLOW',case=False)]
    grand_total_yellow = yellow_df['Sale Amount Actual'].sum()
    total_yellow_profit_loss = second_merge_table[second_merge_table['Dimension1'] == 'YEL GEN']['Profit & Loss'].values[0]
    yellow_df['amnt'] = (yellow_df['Sale Amount Actual'] * total_yellow_profit_loss) / grand_total_yellow
    grand_total_yellow_row = pd.DataFrame({'Department': ['Grand Total'], 'Sale Amount Actual': [grand_total_yellow], 'amnt': [total_yellow_profit_loss]})
    yellow_df = pd.concat([yellow_df, grand_total_yellow_row], ignore_index=True)
    # yellow_df = yellow_df.append({'Department': 'Grand Total', 'Sale Amount Actual':grand_total_yellow}, ignore_index=True)

                                                
    ###########################################################################################################################
                                                

    # Create an in-memory output file
    output_file = io.BytesIO()

    # Create a writer object using the openpyxl engine
    with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
        # Write the dataframes to the Excel file
        first_merged.to_excel(writer, sheet_name='Detailed_Ledger', index=False)

        num_columns = len(Dimension_department.columns)
        second_merge_table.to_excel(writer, sheet_name='Detailed_Ledger', startcol=num_columns + 3, index=False)

        start_row_raw_file = second_merge_table.shape[0] + 4
        red_df.to_excel(writer, sheet_name='Detailed_Ledger', startcol=num_columns + 3, startrow=start_row_raw_file, index=False)

        start_row_yellow = start_row_raw_file + red_df.shape[0] + 2
        yellow_df.to_excel(writer, sheet_name='Detailed_Ledger', startcol=num_columns + 3, startrow=start_row_yellow, index=False)

    # At this point, ExcelWriter has finished writing, but the file is still in the buffer.
    output_file.seek(0)  # Reset buffer position

    # Load the workbook to apply formatting
    wb = load_workbook(output_file)
    ws = wb['Detailed_Ledger']

    # Define the styles for formatting
    bold_font = Font(bold=True)
    orange_fill = PatternFill(start_color="F4B084", end_color="F4B084", fill_type="solid")
    border_style = Border(
        left=Side(border_style='thin'),
        right=Side(border_style='thin'),
        top=Side(border_style='thin'),
        bottom=Side(border_style='thin')
    )

    # Comma style for numbers with thousands separator
    def apply_comma_format(sheet):
        for row in sheet.iter_rows(min_row=2, min_col=1):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'
                    cell.border = border_style

    # Auto-fit column width
    def autofit_columns(sheet):
        for column in sheet.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

    # Apply formatting for 'Grand Total'
    def format_grand_total(sheet):
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                if cell.value == 'Grand Total':
                    cell.font = bold_font
                    cell.fill = orange_fill
                    cell.border = border_style

    # Apply formatting functions
    apply_comma_format(ws)
    autofit_columns(ws)
    format_grand_total(ws)

    # Save the workbook back to the in-memory file
    output_file.seek(0)  # Reset buffer position before saving
    wb.save(output_file)

    # Reset the buffer for downloading
    output_file.seek(0)
    
    # Store the generated file in the Flask app config
    current_app.config['EXCEL_FILE'] = {
        'name': 'Ledger_MIS_Report.xlsx',
        'data': output_file.getvalue()}

    # Redirect to the processing page
    return redirect(url_for('processing'))

@ledger.route('/processing')
def processing():
    return render_template('processing.html')

@ledger.route('/download')
def download_report():
# Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])

if __name__ == '__main__':
    ledger.run(debug=True)