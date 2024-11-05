from flask import Flask,Blueprint,current_app, render_template, request, send_file, redirect, url_for
import io
import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter

department = Blueprint('department', __name__)

# Function to apply borders to the worksheet
def apply_borders(worksheet):
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                cell.border = thin_border

# Function to auto-fit columns and apply formatting
def auto_fit_columns(worksheet):
    for col in worksheet.columns:
        max_length = max(len(str(worksheet.cell(row=1, column=col[0].column).value)),
                         max(len(str(cell.value)) for cell in col if cell.value is not None))
        adjusted_width = (max_length + 2) * 1.2
        worksheet.column_dimensions[get_column_letter(col[0].column)].width = adjusted_width

# Function to format values with thousand separators
def format_thousand_separator(worksheet):
    for row in worksheet.iter_rows(min_row=2):
        for cell in row:
            if isinstance(cell.value, (int, float)):  # Apply formatting to numeric cells only
                cell.number_format = '#,##0'

# Function to apply custom formatting to 'Total Sales' row
def format_total_sales_row(worksheet):
    fill = PatternFill(start_color="B7DEE8", end_color="B7DEE8", fill_type="solid")
    font = Font(bold=True)
    for row in worksheet.iter_rows():
        for cell in row:
            if cell.value == 'Total Sales':  # Check if the cell value is 'Total Sales'
                for cell in row:
                    cell.fill = fill
                    cell.font = font


@department.route('/')
def index():
    return render_template('test.html')

@department.route('/upload', methods=['POST'])
def upload_files():

    trial = request.files['trial']
    sales = request.files['sales']

    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\Allocation (Corporate Department).xlsx'
    # Check if an allocation file was provided by the user
    allocation_file = request.files.get('allocation')
    if allocation_file:
        # If the user provided an allocation file, delete the existing default file
        if os.path.exists(default_allocation_path):
            os.remove(default_allocation_path)  # Delete the existing default allocation file
        allocation_file.save(default_allocation_path)  # Save the new allocation file
        allocation = pd.read_excel(allocation_file)
    else:
        # If no allocation file is provided, use the default allocation file
        allocation = pd.read_excel(default_allocation_path)

    # Path to the default last year sales file
    default_lastyr_path = r'D:\2MJIPL_COIMBATORE\Last Year Sales Data.xlsx'
    # Check if a last year sales file was provided by the user
    lastyr_file = request.files.get('lastsales')
    if lastyr_file:
        # If the user provided a last year sales, delete the existing default file
        if os.path.exists(default_lastyr_path):
            os.remove(default_lastyr_path)  # Delete the existing default_lastyr_path
        lastyr_file.save(default_lastyr_path)  # Save the new last year sales file
        last_yr = pd.read_excel(lastyr_file)
    else:
        # If no allocation file is provided, use the default allocation file
        last_yr = pd.read_excel(default_lastyr_path)

# Read the uploaded Excel files

    trial_file = pd.read_excel(trial, sheet_name = 'detailed Ledger Mis Repor...')
    sales_file = pd.read_excel(sales, sheet_name='Default Layout')
    # last_yr = pd.read_excel(lastsales)

    month_order = ['April-24', 'May-24', 'June-24', 'July-24', 'August-24', 'September-24',
            'October-24', 'November-24', 'December-24', 'January-25', 'February-25', 'March-25' ]

    filter_corporate = trial_file[trial_file['Branch'] == 'Corporate Branch']

    filter_again_added =['Tea Procurement','Red General Inst']
    filter_corporate_new =filter_corporate[~filter_corporate['Departments'].isin(filter_again_added)]

    rows_to_remove = ['Africa Project', 'Ghana', 'Ghana Trade', 'Non Corporate', 'Russia Trade', 'Specialty']
    filter_department =filter_corporate_new[~filter_corporate_new['Departments'].isin(rows_to_remove)]
    modify_df = filter_department[(filter_department['Ledger Groups'] != 'Balancing Accounts')]


    # not_taken_al = ['Non Production' , 'To be ignored']
    # allocation = allocation[~allocation['Allocation'].isin(not_taken_al)]
    #

    rows_to_dlt= ['Bank Guarantee charges','Bank Charges (Gst)','Bank Charges Misc','Bank Charges Import /Ott','Bank Charges Realisation',
    'Foreign Bank Charges','Forex Conversion Chgs (Gst)','Interest On Epc-Icici Bank Ltd.','Interest On Hdfc Cc',
    'Interest On Icici Bank Cash Credit Account','Interest On Icici Bank-Wcdl','Interest On Working Capital Term Loan',
    'Interest paid Misc','Notional Reimbursement Interest On Capital-Income','Notional Reimbursement Of Rent of Factory Premises',
    'Notional Reimbursement Of Rent Of Factory Premises-Income','Notional Transfer Of Salary','Customs Duty Drawback Claims',
    'Preshipment Regular Interest','Direct Cost Applied','Notional Transfer of expenses','Profit & Loss on Sale of Fixed Assets',
    'Notional Transfer of expenses','Notional Transfer of Finished Goods at Target Pric','R&M Corporate','Notional Transfer of expenses ']



    clean_df = modify_df[~modify_df['Ledger Name'].isin(rows_to_dlt)]
    clean_df.loc[: ,'Month'] = clean_df['Voucher Date'].dt.strftime('%B-%y')

    #######################################################################################################################################################

    allocation_clean =allocation.drop_duplicates(subset=['Branch','Departments','Ledger Name','Allocation'])
    raw_file_df =pd.merge(clean_df,allocation_clean, on=['Branch','Departments','Ledger Name'], how = 'left')
    raw_file_df = raw_file_df[['Branch', 'Departments','Ledger Groups', 'Ledger Name','Month','Profit & Loss','Allocation']]



    corporate_df = clean_df[clean_df['Departments'] == 'Corporate']
    corporate_df_groupby =corporate_df.groupby(["Ledger Name","Month"])["Profit & Loss"].sum().reset_index()
    corporate_df_groupby

    #######################################################################################################################################################


    allocation_clean =allocation.drop_duplicates(subset=['Branch','Departments','Ledger Name','Allocation'])
    master_with_allocation = corporate_df_groupby.merge(allocation_clean, on='Ledger Name', how = 'left')


    mdf =master_with_allocation.pivot_table(index='Allocation', columns='Month', values='Profit & Loss', aggfunc='sum',fill_value=0)



    mdf_new = mdf.reindex(columns=month_order)
    mdf_new['April2024 to March2025'] = mdf_new.loc[:, 'April-24':'March-25'].sum(axis=1)

    ######################################################################################################################################################

    pivot_department = clean_df.pivot_table(index='Departments', columns=['Month'],
                                    values='Profit & Loss', aggfunc='sum',fill_value=0)

    Department_table = pivot_department.reindex(columns=month_order)
    Department_table['April2024 to March2025'] =Department_table.loc[:, 'April-24':'March-25'].sum(axis=1)
    total_row = Department_table.sum(axis=0)
    Department_table.loc['Actual Corporate Cost incurred'] = total_row
    Department_table

    #####################################################################################################################################################

    sale_df = sales_file[sales_file['G/L Ac-Description'].isin(['3000100000 - Sales Tea Domestic', '3000040000 - Sales Tea Export'])]
    sale_df.loc[: ,'Month'] = sale_df['Invoice Date'].dt.strftime('%B-%y')

    pd.options.display.float_format = '{:.2f}'.format

    sale_table = sale_df.pivot_table(index='G/L Ac-Description', columns=['Month'],
                                    values='Sale Amount Actual', aggfunc='sum',fill_value=0)

    sale_table = sale_table.reindex(columns=month_order)
    sale_table['April2024 to March2025'] = sale_table.loc[:, 'April-24':'March-25'].sum(axis=1)
    total_row = sale_table.sum(axis=0)

    sale_table.loc['Total Sales'] = total_row
    pd.set_option('display.float_format', '{:.2f}'.format)

    #####################################################################################################################################################

    merged_table = pd.concat([Department_table, sale_table])

    actual_corporate_cost_rate = (merged_table.loc['Actual Corporate Cost incurred'] / merged_table.loc['Total Sales']) * 100

    formatted_rates = ["{:.2f}%".format(value) for value in actual_corporate_cost_rate]

    merged_table.loc['Actual corporate cost rate'] = formatted_rates

    merged_table.rename(index={'3000040000 - Sales Tea Export': 'Export Sales', '3000100000 - Sales Tea Domestic': 'Domestic Sales'}, inplace=True)
    merged_table.index.name = 'Departments'



    ###################################################
    last_yr_YTD = last_yr[["YTD 23-24"]]
    branch_to_am = last_yr_YTD["YTD 23-24"].tolist()
    branch_to_am_series = pd.Series(branch_to_am, index=merged_table.index)
    merged_table["YTD  Apr'23-Apr'24"] = branch_to_am_series
    #####################################################

    error_df =raw_file_df[['Branch','Departments','Ledger Name', 'Allocation']]
    error_file = error_df[(error_df['Departments'] == 'Corporate') & (error_df['Allocation'].isna())]
    error_file = error_file.drop_duplicates()

    # Create an in-memory output file
    output = io.BytesIO()

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        merged_table.round().to_excel(writer, sheet_name='Departments', index=True)
        mdf_new.round().to_excel(writer, sheet_name='Corporate', index=True)
        raw_file_df.to_excel(writer, sheet_name='RawFile', index=False)
        error_file.to_excel(writer, sheet_name='ERROR_Rows', index=False)

        # Access the worksheets and apply formatting
        workbook = writer.book
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            auto_fit_columns(worksheet)
            format_thousand_separator(worksheet)
            apply_borders(worksheet)
            format_total_sales_row(worksheet)

    output.seek(0)  # Reset file pointer to the beginning of the file

    current_app.config['EXCEL_FILE'] = {
        'name': 'Department_corporate.xlsx',
        'data': output.getvalue()}


    # Redirect to the processing page
    return redirect(url_for('processing'))

@department.route('/processing')
def processing():
    return render_template('processing.html')

@department.route('/download')
def download_report():
# Fetch the in-memory file
    file_info = current_app.config.get('EXCEL_FILE')
    if not file_info:
        return "No file found", 404
    return send_file(io.BytesIO(file_info['data']), as_attachment=True, download_name=file_info['name'])

@department.route('/department-master', methods=['GET'])
def department_master():
    # Path to the default allocation file
    default_allocation_path = r'D:\2MJIPL_COIMBATORE\Allocation (Corporate Department).xlsx'
    return send_file(default_allocation_path, as_attachment=True)

@department.route('/department-lastyr', methods=['GET'])
def department_lastyr():
    # Path to the default lastyear sales file
    default_lastyr_path = r'D:\2MJIPL_COIMBATORE\Last Year Sales Data.xlsx' 
    return send_file(default_lastyr_path, as_attachment=True)

if __name__ == '__main__':
    department.run(debug=True)